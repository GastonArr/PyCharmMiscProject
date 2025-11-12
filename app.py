import streamlit as st
from openpyxl import load_workbook, Workbook
import os
import datetime
import direcciones          # módulo externo para pantalla de direcciones streamlit run app.py
import Robos_Hurtos         # subflujo para delitos Robos/Hurtos
import otros                # subflujo para Lesiones / Desaparición
import agenda_delitos       # gestión de almanaque de delitos asignados
from login import render_login, render_user_header

# ===========================
# Config de rutas (en el repo)
# ===========================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_DIR = os.path.join(BASE_DIR, "excel")  # carpeta local en el repo
os.makedirs(EXCEL_DIR, exist_ok=True)

# Bases SIN extensión (usamos resolve_excel_path para elegir .xlsm/.xlsx/.xls)
excel_base_comisaria_14 = os.path.join(EXCEL_DIR, "comisaria 14")
excel_base_comisaria_15 = os.path.join(EXCEL_DIR, "comisaria 15")
excel_base_comisaria_6  = os.path.join(EXCEL_DIR, "comisaria 6")
excel_base_comisaria_42 = os.path.join(EXCEL_DIR, "comisaria 42")
excel_base_comisaria_9  = os.path.join(EXCEL_DIR, "comisaria 9")
excel_base_cenaf_4      = os.path.join(EXCEL_DIR, "CENAF 4")

# ---------------------------
# Utilidades
# ---------------------------

def is_xlsm(path: str) -> bool:
    return path.lower().endswith(".xlsm")

def resolve_excel_path(base_without_ext: str) -> str:
    """
    Devuelve el archivo existente entre: .xlsm, .xlsx, .xls (en ese orden).
    Si no existe ninguno, devuelve base + '.xlsm'.
    """
    for ext in (".xlsm", ".xlsx", ".xls"):
        cand = base_without_ext + ext
        if os.path.exists(cand):
            return cand
    return base_without_ext + ".xlsm"

def asegurar_excel(path: str):
    """
    Crea el archivo si no existe (xlsx o xlsm según la extensión).
    No crea macros; si el archivo ya tiene macros, se conservarán con keep_vba=True al cargar/guardar.
    """
    carpeta = os.path.dirname(path)
    if carpeta and not os.path.exists(carpeta):
        os.makedirs(carpeta, exist_ok=True)
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Hoja1"
        wb.save(path)

def cargar_libro(path: str):
    """
    Carga el libro respetando macros si es .xlsm.
    """
    asegurar_excel(path)
    return load_workbook(path, keep_vba=is_xlsm(path))

def obtener_siguiente_fila_por_fecha(path: str, col_fecha: str = "C") -> int:
    """
    Busca la primera fila vacía en la columna de fecha (C por defecto) a partir de la fila 3.
    """
    wb = cargar_libro(path)
    ws = wb.active
    fila = 3
    while True:
        val = ws[f"{col_fecha}{fila}"].value
        if val is None or (isinstance(val, str) and val.strip() == ""):
            return fila
        fila += 1

def unwrap_quotes(v):
    if not isinstance(v, str):
        return v
    if len(v) >= 2 and ((v[0] == v[-1] == '"') or (v[0] == v[-1] == "'")):
        return v[1:-1]
    return v

def fecha_a_texto_curvo(fecha: datetime.date) -> str:
    if not isinstance(fecha, datetime.date):
        return None
    return fecha.strftime("%d/%m/%y")

def escribir_registro(path: str, fila: int, hecho, delito,
                      actuacion, fecha_denuncia_txt, fecha_hecho_txt,
                      hora_hecho_txt, hora_fin_txt, preventivo_txt,
                      denunciante_txt, motivo_txt):
    """
    Escribe datos en: C,D,E,F,H,Q,AF,BL,X,R (sin tocar A).
    """
    wb = cargar_libro(path)
    ws = wb.active
    try:
        ws[f"C{fila}"].value  = fecha_denuncia_txt
        ws[f"D{fila}"].value  = fecha_hecho_txt
        ws[f"E{fila}"].value  = hora_hecho_txt
        ws[f"F{fila}"].value  = hora_fin_txt
        ws[f"H{fila}"].value  = unwrap_quotes(preventivo_txt)
        ws[f"Q{fila}"].value  = unwrap_quotes(denunciante_txt)
        ws[f"AF{fila}"].value = unwrap_quotes(motivo_txt)
        ws[f"BL{fila}"].value = unwrap_quotes(hecho)
        ws[f"X{fila}"].value  = unwrap_quotes(delito)
        ws[f"R{fila}"].value  = unwrap_quotes(actuacion)
        wb.save(path)
        return True
    except PermissionError:
        st.error("⚠️ No se pudo guardar porque el archivo está abierto en Excel con bloqueo de escritura. Cerrá el archivo y probá de nuevo.")
        return False
    except Exception as e:
        st.error(f"⚠️ No se pudo escribir en {path}: {e}")
        return False

def mostrar_hecho():
    if st.session_state.get("hecho"):
        st.subheader("Hecho ingresado para referencia:")
        st.write(st.session_state.hecho)

def excel_path_por_comisaria(nombre_comisaria: str) -> str:
    """
    Devuelve el path del Excel (con extensión) según la comisaría seleccionada,
    resolviendo la extensión existente o creando .xlsm si no hay.
    """
    if nombre_comisaria == "Comisaria 14":
        base = excel_base_comisaria_14
    elif nombre_comisaria == "Comisaria 15":
        base = excel_base_comisaria_15
    elif nombre_comisaria == "Comisaria 6":
        base = excel_base_comisaria_6
    elif nombre_comisaria == "Comisaria 42":
        base = excel_base_comisaria_42
    elif nombre_comisaria == "Comisaria 9":
        base = excel_base_comisaria_9
    else:  # "CENAF 4"
        base = excel_base_cenaf_4
    return resolve_excel_path(base)

# ---------------------------
# Estado de sesión
# ---------------------------

def _init_state():
    d = st.session_state
    d.setdefault("authenticated", False)
    d.setdefault("username", None)
    d.setdefault("allowed_comisarias", None)
    d.setdefault("step", 1)
    d.setdefault("comisaria", None)
    d.setdefault("delito", None)
    d.setdefault("agenda_fecha", None)
    d.setdefault("hecho", None)
    d.setdefault("actuacion", None)
    d.setdefault("excel_path", None)
    d.setdefault("fila", None)
    d.setdefault("fecha_denuncia", None)
    d.setdefault("fecha_hecho", None)
    d.setdefault("hora_hecho", None)
    d.setdefault("hora_fin", None)
    d.setdefault("preventivo", "")
    d.setdefault("denunciante", "")
    d.setdefault("motivo", "")
    # Subflujos
    d.setdefault("rh_done", False)
    d.setdefault("rh_preview", None)       # resumen Robos/Hurtos
    d.setdefault("others_done", False)
    d.setdefault("others_preview", None)   # resumen Otros (Lesiones/Desap)
    d.setdefault("direcciones_preview", None)  # resumen Direcciones

_init_state()

# Bloquear acceso si no está autenticado
if not st.session_state.authenticated:
    render_login()
    st.stop()

render_user_header()

# Validar que la comisaría seleccionada (si existe) sea permitida
allowed_comisarias = st.session_state.allowed_comisarias or []

# Panel de administración del almanaque (solo usuarios habilitados)
agenda_delitos.render_admin_agenda(st.session_state.username, allowed_comisarias)

if st.session_state.comisaria and st.session_state.comisaria not in allowed_comisarias:
    st.session_state.comisaria = None
    st.session_state.step = 1

# ---------------------------
# UI por pasos
# ---------------------------

if st.session_state.step == 1:
    st.title("CARGA DE SNIC")
    st.subheader("Seleccione la comisaría en la que desea trabajar:")

    opciones_comisaria = allowed_comisarias
    if not opciones_comisaria:
        st.error("Su usuario no tiene comisarías asignadas. Contacte al administrador del sistema.")
        st.stop()

    if st.session_state.comisaria in opciones_comisaria:
        index_default = opciones_comisaria.index(st.session_state.comisaria)
    else:
        index_default = 0

    comisaria = st.selectbox(
        "Seleccione la comisaría",
        opciones_comisaria,
        index=index_default if opciones_comisaria else 0,
    )

    excel_path_preview = excel_path_por_comisaria(comisaria)
    asegurar_excel(excel_path_preview)
    fila_objetivo = obtener_siguiente_fila_por_fecha(excel_path_preview, col_fecha="C")
    planilla_llena = fila_objetivo >= 103

    if planilla_llena:
        st.error("PLANILLA COMPLETA POR FAVOR RENUEVE.")

    # --- Botón Descargar Excel + Uploader con validación de nombre ---
    col_dl, col_next = st.columns([1,1])

    with col_dl:
        # Detectar MIME según extensión
        _ext = os.path.splitext(excel_path_preview)[1].lower()
        if _ext == ".xlsm":
            mime = "application/vnd.ms-excel.sheet.macroEnabled.12"
        elif _ext == ".xlsx":
            mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:  # .xls u otro
            mime = "application/vnd.ms-excel"

        try:
            with open(excel_path_preview, "rb") as f:
                excel_bytes = f.read()
            st.download_button(
                label="📥 Descargar Excel",
                data=excel_bytes,
                file_name=os.path.basename(excel_path_preview),
                mime=mime,
                use_container_width=True,
            )
        except Exception as e:
            st.caption(f"⚠️ No se pudo preparar la descarga: {e}")

        # Uploader: exige que el nombre del archivo subido sea EXACTAMENTE el esperado
        st.markdown("— o —")
        expected_name = os.path.basename(excel_path_preview)
        uploaded = st.file_uploader(
            f"Subir/Reemplazar Excel (nombre requerido: **{expected_name}**)",
            type=["xlsm", "xlsx", "xls"],
            key="uploader_excel"
        )
        if uploaded is not None:
            if uploaded.name != expected_name:
                st.error(f"El archivo debe llamarse **{expected_name}**. Renombralo y volvé a subirlo para evitar conflictos.")
            else:
                # Guardar lo subido SOBRE el archivo target
                try:
                    with open(excel_path_preview, "wb") as f:
                        f.write(uploaded.getbuffer())
                    st.success(f"Se reemplazó el Excel de {comisaria}: {expected_name}")
                except Exception as e:
                    st.error(f"No se pudo guardar el archivo: {e}")

    with col_next:
        if st.button("Siguiente", use_container_width=True):
            if comisaria not in allowed_comisarias:
                st.error("No tiene permisos para trabajar con la comisaría seleccionada.")
                st.stop()
            if planilla_llena:
                st.error("PLANILLA COMPLETA POR FAVOR RENUEVE.")
            else:
                st.session_state.comisaria = comisaria
                st.session_state.excel_path = excel_path_preview
                st.session_state.fila = fila_objetivo
                st.session_state.agenda_fecha = None

                # reset subflujos
                for k in ("rh_cache", "rh_vict_rows", "rh_sex_rows", "rh_step"):
                    st.session_state.pop(k, None)

                st.session_state.rh_done = False
                st.session_state.rh_preview = None
                st.session_state.others_done = False
                st.session_state.others_preview = None
                st.session_state.direcciones_preview = None

                st.session_state.step = 2
                st.rerun()

elif st.session_state.step == 2:
    mostrar_hecho()
    st.subheader(f"Usted seleccionó la {st.session_state.comisaria}")
    st.caption(f"Próximo registro: fila {st.session_state.fila}")

    hecho = st.text_area(
        "Indique el hecho (copie tal cual del MEMORANDUM):",
        height=150,
        value=st.session_state.hecho or ""
    )
    preventivo = st.text_input(
        "Ingrese el número de preventivo (máx 20 caracteres)",
        value=st.session_state.preventivo or "",
        max_chars=20
    )

    col1, col2 = st.columns(2)
    with col1:
        if st.button("Volver"):
            st.session_state.agenda_fecha = None
            st.session_state.step = 1
            st.rerun()
    with col2:
        if st.button("Siguiente"):
            if not hecho or hecho.strip() == "":
                st.warning("Por favor, ingrese el hecho antes de continuar.")
                st.stop()
            if not preventivo or preventivo.strip() == "":
                st.warning("Por favor, ingrese el número de preventivo.")
                st.stop()
            st.session_state.hecho = hecho
            st.session_state.preventivo = preventivo
            st.session_state.step = 3
            st.rerun()

elif st.session_state.step == 3:
    mostrar_hecho()
    st.subheader("Seleccione el día y el delito asignado")

    comisaria_actual = st.session_state.comisaria
    fecha_agenda, delitos_pendientes, mensaje_agenda = agenda_delitos.render_selector_comisaria(comisaria_actual)

    if mensaje_agenda:
        st.warning(mensaje_agenda)
        st.stop()

    if fecha_agenda is None:
        st.warning("Seleccione un día válido para continuar.")
        st.stop()

    if not delitos_pendientes:
        st.info("No hay delitos pendientes para el día seleccionado. Consulte al administrador para nuevas asignaciones.")
        st.stop()

    opciones_delito = list(delitos_pendientes.keys())

    def _format_delito(nombre: str) -> str:
        info_delito = delitos_pendientes.get(nombre, {})
        restantes = info_delito.get("restantes", 0)
        planificados = info_delito.get("plan", 0)
        return f"{nombre.strip()} — restantes {restantes} de {planificados}"

    if st.session_state.delito in opciones_delito:
        index_default = opciones_delito.index(st.session_state.delito)
    else:
        index_default = 0

    delito = st.selectbox(
        "Seleccione el delito",
        opciones_delito,
        index=index_default,
        format_func=_format_delito,
    )
    st.caption("Solo se muestran los delitos asignados por el administrador para el día elegido.")

    denunciante = st.text_input(
        "Indique el nombre y apellido del denunciante o víctima",
        value=st.session_state.denunciante or "",
        max_chars=100
    )
    motivos = [
        "DENUNCIA PARTICULAR",
        "INTERVENCIÓN POLICIAL",
        "ORDEN JUDICIAL",
        "OTROS/ NO CONSTA",
    ]
    motivo_sel = st.selectbox(
        "Motivo que origina el registro del hecho",
        motivos,
        index=(motivos.index(st.session_state.motivo) if st.session_state.motivo in motivos else 0),
    )

    col1, col2 = st.columns(2)
    with col1:
        if st.button("Volver"):
            st.session_state.step = 2
            st.rerun()
    with col2:
        if st.button("Siguiente"):
            if not denunciante or denunciante.strip() == "":
                st.warning("Por favor, ingrese el nombre y apellido del denunciante o víctima.")
                st.stop()

            fecha_validacion = st.session_state.get("agenda_fecha")
            if not isinstance(fecha_validacion, datetime.date):
                st.warning("Seleccione un día válido antes de continuar.")
                st.stop()

            ok_delito, msg_delito = agenda_delitos.puede_cargar_delito(
                st.session_state.comisaria,
                fecha_validacion,
                delito,
            )
            if not ok_delito:
                st.warning(msg_delito or "El delito seleccionado no está disponible para su carga.")
                st.stop()

            # NUEVO: si cambió el delito, limpiar previews/flags/caches de subflujos
            prev_delito_norm = (st.session_state.delito or "").strip() if st.session_state.get("delito") else None
            nuevo_delito_norm = (delito or "").strip()
            if prev_delito_norm and prev_delito_norm != nuevo_delito_norm:
                # Robos/Hurtos
                for k in ("rh_done", "rh_preview", "rh_cache", "rh_vict_rows", "rh_sex_rows", "rh_step"):
                    if k in st.session_state:
                        del st.session_state[k]
                # Otros
                for k in ("others_done", "others_preview", "others_vict_rows", "others_step"):
                    if k in st.session_state:
                        del st.session_state[k]
                # Direcciones NO se toca

            st.session_state.delito = delito
            st.session_state.denunciante = denunciante
            st.session_state.motivo = motivo_sel
            st.session_state.step = 4
            st.rerun()

elif st.session_state.step == 4:
    mostrar_hecho()
    st.subheader("Indique el tipo de actuación")

    delito_actual = (st.session_state.delito or "").strip()
    if delito_actual == "A CONSIDERACION":
        opciones_actuacion = ["A CONSIDERACION"]
        st.session_state.actuacion = "A CONSIDERACION"
        st.caption("Por haber seleccionado el delito 'A CONSIDERACION', la actuación queda fijada en 'A CONSIDERACION'.")
    else:
        opciones_actuacion = ["CONVENCIONAL", "ABREVIADA"]
        if st.session_state.actuacion not in opciones_actuacion:
            st.session_state.actuacion = "CONVENCIONAL"

    actuacion = st.radio(
        "Seleccione una opción",
        opciones_actuacion,
        index=(opciones_actuacion.index(st.session_state.actuacion)
               if st.session_state.actuacion in opciones_actuacion else 0),
        key="actuacion_radio"
    )
    st.session_state.actuacion = actuacion

    st.subheader("Ingrese el día y hora")

    st.markdown("**Fecha de denuncia/intervención:**")
    fecha_denuncia = st.date_input(
        "Fecha de denuncia/intervención",
        value=st.session_state.fecha_denuncia or datetime.date.today(),
        key="fecha_denuncia_input"
    )
    st.session_state.fecha_denuncia = fecha_denuncia

    st.markdown("**Fecha del hecho:**")
    fecha_hecho = st.date_input(
        "Fecha del hecho",
        value=st.session_state.fecha_hecho or datetime.date.today(),
        key="fecha_hecho_input"
    )
    st.session_state.fecha_hecho = fecha_hecho

    opciones_hora = [f"{h:02d}:00" for h in range(24)]
    opciones_hora_con_ind = ["INDETERMINADO"] + opciones_hora

    col_h1, col_h2 = st.columns(2)
    with col_h1:
        st.markdown("**Horario del hecho:**")
        hora_hecho = st.selectbox(
            "Horario del hecho",
            opciones_hora_con_ind,
            index=(opciones_hora_con_ind.index(st.session_state.hora_hecho)
                   if st.session_state.hora_hecho in opciones_hora_con_ind else 0),
            key="hora_hecho_select"
        )
        st.session_state.hora_hecho = hora_hecho

    with col_h2:
        st.markdown("**Horario fin del hecho:**")
        hora_fin = st.selectbox(
            "Horario fin del hecho",
            opciones_hora_con_ind,
            index=(opciones_hora_con_ind.index(st.session_state.hora_fin)
                   if st.session_state.hora_fin in opciones_hora_con_ind else 0),
            key="hora_fin_select"
        )
        st.session_state.hora_fin = hora_fin

    col1, col2 = st.columns(2)
    with col1:
        if st.button("Volver"):
            st.session_state.step = 3
            st.rerun()
    with col2:
        if st.button("Siguiente"):
            st.session_state.step = 5
            st.rerun()

elif st.session_state.step == 5:
    # Direcciones SIEMPRE
    mostrar_hecho()
    direcciones.render_direcciones_ui(
        st.session_state.excel_path,
        st.session_state.fila,
        st.session_state.comisaria
    )
    # La función interna maneja volver (step=4) y continuar (step=6).

elif st.session_state.step == 6:
    # ===========================
    # NUEVO: normalizar y limpiar previews que no correspondan al delito actual
    # ===========================
    delito_now_norm = (st.session_state.delito or "").strip()

    delitos_rh = {
        "ROBO SIMPLE ",
        "HURTO SIMPLE ",
        "ROBO AGRAVADO POR EL USO DE ARMA DE ARMA BLANCA",
        "ROBO AGRAVADO POR EL USO DE ARMA DE FUEGO",
        "ROBO ABIGEATO ",
        "ROBO AGRAVADO POR LESION ",
    }
    delitos_rh_norm = {d.strip() for d in delitos_rh}

    delitos_otros = {
        "LESIONES GRAVES", "LESIONES LEVES", "LESIONES GRAVISIMAS", "OTROS DELITOS CONTRA LA LIBERTAD INDIVIDUAL",
        "DESAPARICION DE PERSONA",
        "ABUSO DE ARMAS CON LESIONES",
        "ABUSO DE ARMAS SIN LESIONES",
        "ABUSO SEXUAL CON ACCESO CARNAL (VIOLACION)",
        "ABUSO SEXUAL SIMPLE",
    }
    delitos_otros_norm = {d.strip() for d in delitos_otros}

    # Si hay previews de un subflujo que ya no aplica, descartarlos
    if st.session_state.get("rh_preview") and (delito_now_norm not in delitos_rh_norm):
        st.session_state.rh_preview = None
        st.session_state.rh_done = False
        for k in ("rh_cache", "rh_vict_rows", "rh_sex_rows", "rh_step"):
            st.session_state.pop(k, None)

    if st.session_state.get("others_preview") and (delito_now_norm not in delitos_otros_norm):
        st.session_state.others_preview = None
        st.session_state.others_done = False

    # --- Subflujo: Lesiones / Desaparición de Persona (otros.py) ---
    if (delito_now_norm in delitos_otros_norm) and not st.session_state.get("others_done", False):
        otros.render(
            excel_path=st.session_state.excel_path,
            fila=st.session_state.fila,
            delito_x3=st.session_state.delito
        )
        st.stop()

    # --- Subflujo: Robos / Hurtos (Robos_Hurtos.py) ---
    if ((st.session_state.delito in delitos_rh) or (delito_now_norm in delitos_rh_norm)) \
       and not st.session_state.get("rh_done", False):
        Robos_Hurtos.render(
            excel_path=st.session_state.excel_path,
            fila=st.session_state.fila,
            delito_x3=st.session_state.delito
        )
        st.stop()

    # ----- Vista previa antes de guardar -----
    mostrar_hecho()
    st.subheader("Vista previa de lo ingresado")

    # Bloque general
    st.markdown("**Datos generales**")
    st.write(f"- Comisaría: {st.session_state.comisaria}")
    st.write(f"- Fila de escritura: {st.session_state.fila}")
    st.write(f"- Delito: {st.session_state.delito}")
    st.write(f"- Actuación: {st.session_state.actuacion}")
    st.write(f"- Fecha denuncia/intervención: {st.session_state.fecha_denuncia}")
    st.write(f"- Fecha del hecho: {st.session_state.fecha_hecho}")
    st.write(f"- Hora del hecho: {st.session_state.hora_hecho}")
    st.write(f"- Hora fin del hecho: {st.session_state.hora_fin}")
    st.write(f"- N° de preventivo: {st.session_state.preventivo}")
    st.write(f"- Denunciante/Víctima: {st.session_state.denunciante}")
    st.write(f"- Motivo: {st.session_state.motivo}")

    # Bloque Lesiones / Desaparición (si hay y corresponde al delito actual)
    otros_preview = st.session_state.get("others_preview")
    if otros_preview and ((st.session_state.delito or "").strip() in delitos_otros_norm):
        st.markdown("---")
        st.markdown("**Datos adicionales (Lesiones / Desaparición)**")
        vict_rows_o = otros_preview.get("vict_rows") or []
        if vict_rows_o:
            st.write("• Víctimas por sexo:")
            for r in vict_rows_o:
                st.write(f"  - {r.get('sexo')}: {r.get('cant')}")
            st.write(f"  - **Total (AO)**: {otros_preview.get('vict_total')}")
        st.write(f"• Vulnerabilidad (AP): {otros_preview.get('vulnerabilidad')}")
        if otros_preview.get("aparecio") is not None:
            st.write(f"• ¿Apareció? (BA): {otros_preview.get('aparecio')}")

        # Botón para editar subflujo Otros
        if st.button("Editar Lesiones/Desaparición"):
            st.session_state.others_done = False
            st.session_state.others_step = 1
            st.rerun()

    # Bloque Robos/Hurtos (si hay y corresponde al delito actual)
    rh_preview = st.session_state.get("rh_preview")
    if rh_preview and ((st.session_state.delito or "").strip() in delitos_rh_norm):
        st.markdown("---")
        st.markdown("**Datos adicionales (Robos/Hurtos)**")
        # Víctimas
        vict_rows = rh_preview.get("vict_rows") or []
        vict_total = rh_preview.get("vict_total")
        if vict_rows:
            st.write("• Víctimas por sexo:")
            for r in vict_rows:
                st.write(f"  - {r.get('sexo')}: {r.get('cant')}")
            st.write(f"  - **Total (AO)**: {vict_total}")
        # Vulnerabilidad, arma
        st.write(f"• Vulnerabilidad: {rh_preview.get('vulnerab')}")
        st.write(f"• Tipo de arma (AR): {rh_preview.get('tipo_arma')}")
        # Inculpados
        if rh_preview.get("inc_sn") == "SI":
            st.write("• Inculpados: SI")
            st.write(f"  - Rango etario: {rh_preview.get('rango_etario')} ({rh_preview.get('cant_rango')})")
            sex_rows = rh_preview.get("sex_rows") or []
            if sex_rows:
                st.write(f"  - Distribución por sexo:")
                for r in sex_rows:
                    st.write(f"    · {r.get('sexo')}: {r.get('cant')}")
        else:
            st.write("• Inculpados: NO")
        # Tipo de lugar y detalle
        st.write(f"• Tipo de lugar (AD): {rh_preview.get('tipo_lugar')}")
        if rh_preview.get("detalle_est"):
            st.write(f"  - Detalle establecimiento (AE): {rh_preview.get('detalle_est')}")
        # Elementos sustraídos y subdetalles
        st.write(f"• Elementos sustraídos (BB): {rh_preview.get('elem')}")
        if rh_preview.get("subcat"):
            st.write(f"  - Subcategoría (BC): {rh_preview.get('subcat')}")
        if rh_preview.get("denom"):
            st.write(f"  - Denominación (BD): {rh_preview.get('denom')}")
        if rh_preview.get("anio") or rh_preview.get("modelo"):
            st.write(f"  - Año (BE) / Modelo (BF): {rh_preview.get('anio')} / {rh_preview.get('modelo')}")
        # Modus y especialidad
        st.write(f"• Modus operandi (BJ): {rh_preview.get('modus')}")
        if rh_preview.get("especialidad"):
            st.write(f"  - Especialidad (BK): {rh_preview.get('especialidad')}")

        # Botón para editar Robos/Hurtos
        if st.button("Editar Robos/Hurtos"):
            st.session_state.rh_done = False
            st.session_state.rh_step = 1
            st.rerun()

    st.markdown("---")
    st.subheader("Guardar y finalizar")

    col1, col2 = st.columns(2)
    with col1:
        if st.button("Atrás"):
            # Reabrir subflujos si ya fueron completados
            if (delito_now_norm in delitos_otros_norm) and st.session_state.get("others_done", False):
                st.session_state.others_done = False
                st.session_state.others_step = 1
                st.rerun()
            elif ((st.session_state.delito in delitos_rh) or (delito_now_norm in delitos_rh_norm)) \
                 and st.session_state.get("rh_done", False):
                st.session_state.rh_done = False
                st.session_state.rh_step = 1
                st.rerun()
            else:
                st.session_state.step = 4
                st.rerun()

    with col2:
        if st.button("Finalizar y guardar ✅"):
            fila = obtener_siguiente_fila_por_fecha(st.session_state.excel_path, col_fecha="C")

            fecha_den_txt = fecha_a_texto_curvo(st.session_state.fecha_denuncia) if st.session_state.fecha_denuncia else None
            fecha_hecho_txt = fecha_a_texto_curvo(st.session_state.fecha_hecho) if st.session_state.fecha_hecho else None
            hora_hecho_txt = st.session_state.hora_hecho or "INDETERMINADO"
            hora_fin_txt   = st.session_state.hora_fin or "INDETERMINADO"

            # === Guardado diferido de Direcciones (si hay preview) ===
            dprev = st.session_state.get("direcciones_preview")
            if dprev:
                try:
                    wb_dir = cargar_libro(st.session_state.excel_path)
                    ws_dir = wb_dir.active
                    C = lambda col: f"{col}{fila}"

                    # I{fila}: ciudad/código según comisaría
                    cc = dprev.get("ciudad_cod")
                    if cc not in (None, ""):
                        ws_dir[C("I")].value = unwrap_quotes(str(cc))

                    # J{fila}: barrio (si 'OTRO', además guarda el texto en K{fila})
                    b = dprev.get("barrio")
                    if b not in (None, ""):
                        ws_dir[C("J")].value = unwrap_quotes(str(b))
                    ob = dprev.get("otro_barrio") or ""
                    if (b == "OTRO") and ob.strip():
                        ws_dir[C("K")].value = unwrap_quotes(ob.strip())

                    # L{fila}: dirección ; M{fila}: altura
                    dir_txt = dprev.get("direccion")
                    if dir_txt not in (None, ""):
                        ws_dir[C("L")].value = unwrap_quotes(str(dir_txt))
                    alt_txt = dprev.get("altura")
                    if alt_txt not in (None, ""):
                        ws_dir[C("M")].value = unwrap_quotes(str(alt_txt))

                    # N{fila}: link de Google Maps (obligatorio)
                    link = (dprev.get("link_maps") or "").strip()
                    if link:
                        ws_dir[C("N")].value = unwrap_quotes(link)

                    wb_dir.save(st.session_state.excel_path)

                except PermissionError:
                    st.error("⚠️ No se pudo guardar Direcciones: el archivo está abierto en Excel.")
                    st.stop()
                except Exception as e:
                    st.error(f"⚠️ Error al guardar Direcciones: {e}")
                    st.stop()

            # === Guardado diferido de Robos/Hurtos (si hay preview y aplica) ===
            rh_preview = st.session_state.get("rh_preview")
            if rh_preview and ((st.session_state.delito or "").strip() in delitos_rh_norm):
                try:
                    wb_rh = cargar_libro(st.session_state.excel_path)
                    ws_rh = wb_rh.active
                    C = lambda col: f"{col}{fila}"

                    # -------- Víctimas (AO total) + por sexo (AG/AH/AI)
                    total_m = total_f = total_nc = 0
                    for r in (rh_preview.get("vict_rows") or []):
                        sexo = (r.get("sexo") or "").strip()
                        try:
                            c = int(str(r.get("cant") or "0").strip())
                        except Exception:
                            c = 0
                        if sexo == "MASCULINO": total_m += c
                        elif sexo == "FEMENINO": total_f += c
                        elif sexo == "NO CONSTA": total_nc += c
                    ao_total = total_m + total_f + total_nc
                    ws_rh[C("AO")].value = ao_total if ao_total else None
                    if total_m: ws_rh[C("AG")].value = total_m
                    if total_f: ws_rh[C("AH")].value = total_f
                    if total_nc: ws_rh[C("AI")].value = total_nc

                    # -------- Vulnerabilidad (AP) y Tipo de arma (AR)
                    v = rh_preview.get("vulnerab")
                    if v not in (None, ""): ws_rh[C("AP")].value = unwrap_quotes(str(v).strip())
                    ta = rh_preview.get("tipo_arma")
                    if ta not in (None, ""): ws_rh[C("AR")].value = unwrap_quotes(str(ta).strip())

                    # -------- Inculpados: SI/NO (AS) + rango (AW/AX/AY/AZ) + sexo (AT/AU/AV)
                    inc_sn = (rh_preview.get("inc_sn") or "").strip()
                    if inc_sn: ws_rh[C("AS")].value = unwrap_quotes(inc_sn)
                    if inc_sn == "SI":
                        rango = rh_preview.get("rango_etario")
                        cant_rango = rh_preview.get("cant_rango")
                        if rango and str(cant_rango).strip() != "":
                            try:
                                cant_num = int(str(cant_rango).strip())
                            except Exception:
                                cant_num = 0
                            if   rango == "Hasta 15 año":     ws_rh[C("AW")].value = cant_num
                            elif rango == "15 a 17 años":     ws_rh[C("AX")].value = cant_num
                            elif rango == "mayor de 18 años": ws_rh[C("AY")].value = cant_num
                            elif rango == "Sin Determinar":   ws_rh[C("AZ")].value = cant_num
                        # Distribución por sexo de inculpados
                        t_m = t_f = t_nc = 0
                        for r in (rh_preview.get("sex_rows") or []):
                            sx = (r.get("sexo") or "").strip()
                            try:
                                c = int(str(r.get("cant") or "0").strip())
                            except Exception:
                                c = 0
                            if sx == "MASCULINO": t_m += c
                            elif sx == "FEMENINO": t_f += c
                            elif sx == "NO CONSTA": t_nc += c
                        if t_m: ws_rh[C("AT")].value = t_m
                        if t_f: ws_rh[C("AU")].value = t_f
                        if t_nc: ws_rh[C("AV")].value = t_nc

                    # -------- Tipo de lugar (AD) + Detalle establecimiento (AE)
                    tl = rh_preview.get("tipo_lugar")
                    if tl not in (None, ""): ws_rh[C("AD")].value = unwrap_quotes(str(tl).strip())
                    de = rh_preview.get("detalle_est")
                    if de not in (None, ""): ws_rh[C("AE")].value = unwrap_quotes(str(de).strip())

                    # -------- Elementos (BB) + Subcat (BC) + Denom (BD) + Año (BE) + Modelo (BF)
                    el = rh_preview.get("elem")
                    if el not in (None, ""): ws_rh[C("BB")].value = unwrap_quotes(str(el).strip())
                    sc = rh_preview.get("subcat")
                    if sc not in (None, ""): ws_rh[C("BC")].value = unwrap_quotes(str(sc).strip())
                    dn = rh_preview.get("denom")
                    if dn not in (None, ""): ws_rh[C("BD")].value = unwrap_quotes(str(dn).strip())
                    if el in ("AUTOMOTOR", "MOTOCICLETA"):
                        an = rh_preview.get("anio")
                        md = rh_preview.get("modelo")
                        if an not in (None, ""): ws_rh[C("BE")].value = unwrap_quotes(str(an).strip())
                        if md not in (None, ""): ws_rh[C("BF")].value = unwrap_quotes(str(md).strip())

                    # -------- Modus (BJ) + Especialidad (BK)
                    mo = rh_preview.get("modus")
                    if mo not in (None, ""): ws_rh[C("BJ")].value = unwrap_quotes(str(mo).strip())
                    es = rh_preview.get("especialidad")
                    if es not in (None, ""): ws_rh[C("BK")].value = unwrap_quotes(str(es).strip())

                    wb_rh.save(st.session_state.excel_path)

                except PermissionError:
                    st.error("⚠️ No se pudo guardar Robos/Hurtos: el archivo está abierto en Excel.")
                    st.stop()
                except Exception as e:
                    st.error(f"⚠️ Error al guardar datos de Robos/Hurtos: {e}")
                    st.stop()

            # === Guardado diferido de Otros (si hay preview y aplica) ===
            oprev = st.session_state.get("others_preview")
            if oprev and ((st.session_state.delito or "").strip() in delitos_otros_norm):
                try:
                    wb_o = cargar_libro(st.session_state.excel_path)
                    ws_o = wb_o.active
                    C = lambda col: f"{col}{fila}"

                    # Limpiar AO/AG/AH/AI por seguridad (si reescriben tras editar)
                    ws_o[C("AO")].value = None
                    ws_o[C("AG")].value = None
                    ws_o[C("AH")].value = None
                    ws_o[C("AI")].value = None

                    # Acumular por sexo (AG/AH/AI) y total (AO)
                    total_m = total_f = total_nc = 0
                    for r in (oprev.get("vict_rows") or []):
                        sexo = (r.get("sexo") or "").strip()
                        try:
                            c = int(str(r.get("cant") or "0").strip())
                        except Exception:
                            c = 0
                        if sexo == "MASCULINO": total_m += c
                        elif sexo == "FEMENINO": total_f += c
                        elif sexo == "NO CONSTA": total_nc += c
                    ao_total = total_m + total_f + total_nc
                    ws_o[C("AO")].value = ao_total if ao_total else None
                    if total_m: ws_o[C("AG")].value = total_m
                    if total_f: ws_o[C("AH")].value = total_f
                    if total_nc: ws_o[C("AI")].value = total_nc

                    # AP Vulnerabilidad
                    vul = oprev.get("vulnerabilidad")
                    if vul not in (None, ""):
                        ws_o[C("AP")].value = unwrap_quotes(str(vul).strip())

                    # BA ¿Apareció? solo si corresponde (Desaparición)
                    if oprev.get("aparecio") is not None:
                        ws_o[C("BA")].value = unwrap_quotes(str(oprev.get("aparecio")).strip())

                    wb_o.save(st.session_state.excel_path)

                except PermissionError:
                    st.error("⚠️ No se pudo guardar Otros: el archivo está abierto en Excel.")
                    st.stop()
                except Exception as e:
                    st.error(f"⚠️ Error al guardar Otros: {e}")
                    st.stop()

            ok = escribir_registro(
                st.session_state.excel_path,
                fila,
                st.session_state.hecho,
                st.session_state.delito,
                st.session_state.actuacion,
                fecha_den_txt,
                fecha_hecho_txt,
                hora_hecho_txt,
                hora_fin_txt,
                st.session_state.preventivo,   # H
                st.session_state.denunciante,  # Q
                st.session_state.motivo        # AF
            )

            if ok:
                agenda_fecha = st.session_state.get("agenda_fecha")
                if isinstance(agenda_fecha, datetime.date):
                    registrado, msg_agenda, restantes = agenda_delitos.registrar_carga_delito(
                        st.session_state.comisaria,
                        agenda_fecha,
                        st.session_state.delito,
                    )
                    if not registrado:
                        st.warning(msg_agenda or "No se pudo actualizar el almanaque del día seleccionado.")
                    elif restantes == 0:
                        st.caption("✔️ Se completó la carga planificada para este delito en el día seleccionado.")

                st.success(f"Datos guardados en {st.session_state.comisaria} (fila {fila}) ✅")
                # Reset total
                st.session_state.step = 1
                st.session_state.hecho = None
                st.session_state.delito = None
                st.session_state.agenda_fecha = None
                st.session_state.actuacion = None
                st.session_state.fila = None
                st.session_state.fecha_denuncia = None
                st.session_state.fecha_hecho = None
                st.session_state.hora_hecho = None
                st.session_state.hora_fin = None
                st.session_state.preventivo = None
                st.session_state.denunciante = None
                st.session_state.motivo = None
                st.session_state.rh_done = False
                st.session_state.rh_preview = None
                st.session_state.others_done = False
                st.session_state.others_preview = None
                st.session_state.direcciones_preview = None
                st.rerun()
            else:
                st.error("Hubo un problema al guardar. Revise el mensaje de error arriba.")