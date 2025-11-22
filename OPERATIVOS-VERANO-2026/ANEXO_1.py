import os
import streamlit as st
from datetime import date, time, timedelta
from typing import List, Dict
from openpyxl import load_workbook

DATA_START_ROW_ANEXO1 = 7  # fila donde empiezan los datos en el Excel

# Rutas por unidad
UNIT_FILE_ANEXO1 = {
    "comisaria 9":  os.path.join("Excel", "ANEXO I DIAGRAMAS OP VERANO DSICCO-comisaria9.xlsx"),
    "comisaria 42": os.path.join("Excel", "ANEXO I DIAGRAMAS OP VERANO DSICCO-comisaria42.xlsx"),
    "DTCCO-PH":     os.path.join("Excel", "ANEXO I DIAGRAMAS OP VERANO DSICCO-DTCCO-PH.xlsx"),
}


# -------------------------------------------------------------------
# Auxiliares de Excel
# -------------------------------------------------------------------
def _get_next_row_and_counter(path: str, data_start_row: int = DATA_START_ROW_ANEXO1):
    """
    Devuelve (wb, ws, next_row, next_number) para escribir en la columna A.
    """
    wb = load_workbook(path)
    ws = wb.active

    last_row = None
    last_number = None

    # Buscar el último número en la columna A
    for row in range(data_start_row, ws.max_row + 1):
        value = ws.cell(row=row, column=1).value
        if value is not None:
            last_row = row
            last_number = value

    if last_row is None:
        next_row = data_start_row
        next_number = 1
    else:
        next_row = last_row + 1
        try:
            next_number = int(last_number) + 1
        except Exception:
            next_number = 1

    return wb, ws, next_row, next_number


def _guardar_todos_anexo1(ruta_excel: str, diagramas: List[Dict]) -> int:
    """Guarda una lista de diagramas en UN archivo Excel."""
    if not diagramas:
        return 0

    wb, ws, start_row, start_num = _get_next_row_and_counter(ruta_excel, DATA_START_ROW_ANEXO1)

    for idx, d in enumerate(diagramas):
        row = start_row + idx
        num = start_num + idx

        unidad = d["unidad"].strip()
        fecha = d["fecha"]
        rec_hum = d["rec_hum"]
        rec_util = d["rec_util"]
        lugar = d["lugar"].strip()
        hora_desde = d["hora_desde"]
        hora_hasta = d["hora_hasta"]

        ws.cell(row=row, column=1).value = num
        ws.cell(row=row, column=2).value = unidad
        ws.cell(row=row, column=3).value = fecha.strftime("%d/%m/%Y")
        ws.cell(row=row, column=4).value = rec_hum
        ws.cell(row=row, column=5).value = rec_util
        ws.cell(row=row, column=6).value = lugar

        texto_horas = f"{hora_desde.strftime('%H:%M')} - {hora_hasta.strftime('%H:%M')}"
        ws.cell(row=row, column=7).value = texto_horas

    try:
        wb.save(ruta_excel)
    except PermissionError:
        st.error(
            f"No se pudo guardar el archivo '{ruta_excel}'. "
            "Verifique que no esté abierto en otra aplicación."
        )
        st.stop()

    return len(diagramas)


def _guardar_por_unidad_anexo1(diagramas: List[Dict]) -> int:
    """
    Agrupa los diagramas por unidad y guarda cada grupo
    en el archivo correspondiente a esa unidad.
    """
    if not diagramas:
        return 0

    # Agrupar por unidad
    por_unidad: Dict[str, List[Dict]] = {}
    for d in diagramas:
        unidad = d["unidad"].strip()
        por_unidad.setdefault(unidad, []).append(d)

    total = 0
    for unidad, lista in por_unidad.items():
        ruta = UNIT_FILE_ANEXO1.get(unidad)
        if not ruta:
            st.error(
                f"No se encontró un archivo de Excel configurado para la unidad '{unidad}'. "
                "No se guardó ningún dato."
            )
            return 0
        total += _guardar_todos_anexo1(ruta, lista)

    return total


# -------------------------------------------------------------------
# Helpers
# -------------------------------------------------------------------
def _diagrama_vacio(vals: Dict) -> bool:
    """Determina si un diagrama está completamente vacío."""
    if vals["lugar"].strip():
        return False
    if vals["rec_hum"] != 0 or vals["rec_util"] != 0:
        return False
    if vals["hora_desde"] != time(0, 0):
        return False
    if vals["hora_hasta"] != time(0, 0):
        return False
    return True


def _leer_diagrama_desde_state(i: int, fecha_diagrama: date) -> Dict:
    """Lee el contenido del diagrama i desde Streamlit."""
    unidad = st.session_state.get(f"unidad_{i}", "")
    rec_hum = st.session_state.get(f"rec_hum_{i}", 0)
    rec_util = st.session_state.get(f"rec_util_{i}", 0)
    lugar = st.session_state.get(f"lugar_{i}", "")
    hora_desde = st.session_state.get(f"hora_desde_{i}", time(0, 0))
    hora_hasta = st.session_state.get(f"hora_hasta_{i}", time(0, 0))

    return {
        "unidad": unidad,
        "fecha": fecha_diagrama,
        "rec_hum": rec_hum,
        "rec_util": rec_util,
        "lugar": lugar,
        "hora_desde": hora_desde,
        "hora_hasta": hora_hasta,
    }


def _eliminar_diagrama(idx: int, cant: int):
    """Elimina un diagrama de la pantalla y recorre hacia arriba."""
    campos = ["unidad", "rec_hum", "rec_util", "lugar", "hora_desde", "hora_hasta"]

    for j in range(idx, cant - 1):
        for c in campos:
            st.session_state[f"{c}_{j}"] = st.session_state.get(f"{c}_{j+1}")

    for c in campos:
        key = f"{c}_{cant-1}"
        if key in st.session_state:
            del st.session_state[key]

    st.session_state["anexo1_cant_diagramas"] = cant - 1
    st.rerun()


# -------------------------------------------------------------------
# Pantalla principal ANEXO I
# -------------------------------------------------------------------
def mostrar_anexo1(ruta_excel: str, fecha_objetivo: date, fecha_default: date) -> bool:
    """
    ruta_excel se ignora ahora (se usan archivos por unidad),
    pero se mantiene el parámetro para compatibilidad.
    """

    fecha_diagrama = fecha_objetivo + timedelta(days=1)
    unidad_actual = st.session_state.get("unidad_actual", "")

    if "anexo1_cant_diagramas" not in st.session_state:
        st.session_state["anexo1_cant_diagramas"] = 1

    cant = st.session_state["anexo1_cant_diagramas"]

    st.title("BIENVENIDO AL SISTEMA DE CARGA OPERATIVOS VERANO")

    st.markdown("### DIAGRAMA DEL OPERATIVO")
    st.write(
        f"Carga correspondiente al día **{fecha_objetivo.strftime('%d/%m/%Y')}** "
        f"(estos son los operativos que se van a realizar el día "
        f"**{fecha_diagrama.strftime('%d/%m/%Y')}**)."
    )

    st.caption(
        "Complete los campos de cada diagrama. La fecha de operativo se fija "
        "automáticamente como el **día siguiente a la fecha de carga**. "
        "Los datos se guardan en el archivo correspondiente a la unidad seleccionada."
    )

    for i in range(cant):
        st.markdown(f"#### Diagrama {i + 1}")
        with st.container(border=True):

            st.session_state[f"unidad_{i}"] = unidad_actual
            st.text_input(
                "Unidad",
                value=unidad_actual,
                key=f"unidad_{i}",
                disabled=True,
            )

            st.date_input(
                "Fecha del operativo",
                value=fecha_diagrama,
                disabled=True,
                key=f"fecha_{i}",
            )

            st.number_input("Recursos humanos", min_value=0, step=1, key=f"rec_hum_{i}")
            st.number_input("Recursos utilitarios", min_value=0, step=1, key=f"rec_util_{i}")
            st.text_input("Lugar", key=f"lugar_{i}")

            col1, col2 = st.columns(2)
            with col1:
                st.time_input("Hora desde", value=time(0, 0), step=timedelta(hours=1), key=f"hora_desde_{i}")
            with col2:
                st.time_input("Hora hasta", value=time(0, 0), step=timedelta(hours=1), key=f"hora_hasta_{i}")

            if cant > 1 and st.button("Eliminar este diagrama", key=f"del_{i}"):
                _eliminar_diagrama(i, cant)

    st.markdown("---")
    col1, col2 = st.columns(2)

    if col1.button("Agregar otro diagrama"):
        st.session_state["anexo1_cant_diagramas"] = cant + 1
        st.rerun()

    finalizado = False

    if col2.button("Guardar y FINALIZAR LA CARGA"):
        errores = []
        diagramas_validos = []
        vacios = []

        for i in range(cant):
            vals = _leer_diagrama_desde_state(i, fecha_diagrama)

            if _diagrama_vacio(vals):
                vacios.append(i + 1)
                continue

            if not vals["unidad"].strip():
                errores.append(f"Diagrama {i+1}: Unidad es obligatoria.")
            if vals["unidad"].strip() not in UNIT_FILE_ANEXO1:
                errores.append(
                    f"Diagrama {i+1}: Unidad '{vals['unidad']}' no tiene archivo configurado."
                )

            if not vals["lugar"].strip():
                errores.append(f"Diagrama {i+1}: Lugar es obligatorio.")
            if vals["rec_hum"] <= 0:
                errores.append(f"Diagrama {i+1}: Recursos humanos > 0.")
            if vals["rec_util"] <= 0:
                errores.append(f"Diagrama {i+1}: Recursos utilitarios > 0.")
            if vals["hora_desde"].minute != 0 or vals["hora_hasta"].minute != 0:
                errores.append(f"Diagrama {i+1}: Horas deben tener minuto 00.")
            if vals["hora_desde"] == time(0, 0) or vals["hora_hasta"] == time(0, 0):
                errores.append(f"Diagrama {i+1}: Las horas no pueden ser 00:00.")

            diagramas_validos.append(vals)

        if vacios:
            errores.append(
                f"Los diagramas {', '.join(map(str, vacios))} están vacíos. "
                "Elimínelos o complételos."
            )

        if not diagramas_validos and not errores:
            errores.append("Debe completar al menos un diagrama.")

        if errores:
            st.error("Errores encontrados:")
            for e in errores:
                st.write("-", e)
        else:
            n = _guardar_por_unidad_anexo1(diagramas_validos)
            if n > 0:
                if n == 1:
                    st.success("Se ha guardado 1 diagrama y se finalizó la carga.")
                else:
                    st.success(f"Se han guardado {n} diagramas y se finalizó la carga.")
                finalizado = True

    return finalizado
