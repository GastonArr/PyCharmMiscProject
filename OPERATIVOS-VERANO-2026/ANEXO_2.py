import os
import streamlit as st
from datetime import date, time, timedelta
from typing import List, Dict
from openpyxl import load_workbook

DATA_START_ROW_ANEXO2 = 7  # fila donde empiezan los datos en el Excel

# Rutas por unidad
UNIT_FILE_ANEXO2 = {
    "comisaria 9":  os.path.join("Excel", "ANEXO II RESULTADOS OP VERANO-comisaria9.xlsx"),
    "comisaria 42": os.path.join("Excel", "ANEXO II RESULTADOS OP VERANO-comisaria42.xlsx"),
    "DTCCO-PH":     os.path.join("Excel", "ANEXO II RESULTADOS OP VERANO-DTCCO-PH.xlsx"),
}


# -------------------------------------------------------------------
# Auxiliares de Excel
# -------------------------------------------------------------------
def _get_next_row_and_counter(path: str, data_start_row: int = DATA_START_ROW_ANEXO2):
    """
    Devuelve (wb, ws, next_row, next_number) para escribir en la columna A.
    """
    wb = load_workbook(path)
    ws = wb.active

    last_row = None
    last_number = None

    for row in range(data_start_row, ws.max_row + 1):
        value = ws.cell(row=row, column=1).value
        if value is not None:
            last_row = row
            last_number = value

    if last_row is None:
        next_row = data_start_row
        next_number = 1
    else:
        next_row = last_row + 1
        try:
            next_number = int(last_number) + 1
        except Exception:
            next_number = 1

    return wb, ws, next_row, next_number


def _guardar_todos_anexo2(ruta_excel: str, resultados: List[Dict]) -> int:
    """Guarda una lista de resultados en UN archivo Excel."""
    if not resultados:
        return 0

    wb, ws, start_row, start_num = _get_next_row_and_counter(
        ruta_excel, DATA_START_ROW_ANEXO2
    )

    for idx, r in enumerate(resultados):
        row = start_row + idx
        num = start_num + idx

        fecha_val = r["fecha"]
        hora_desde = r["hora_desde"]
        hora_hasta = r["hora_hasta"]
        tipo_op = r["tipo_op"]
        lugar = r["lugar"]
        pers_ident = r["pers_ident"]
        pers_asist = r["pers_asist"]
        delito_prop = r["delito_prop"]
        delito_pers = r["delito_pers"]
        delito_otro = r["delito_otro"]
        lesionados = r["lesionados"]
        dem_aa = r["dem_aa"]
        dem_av_hecho = r["dem_av_hecho"]
        dem_infraganti = r["dem_infraganti"]
        dem_contrav = r["dem_contrav"]
        rec_hum = r["rec_hum"]
        rec_mat = r["rec_mat"]
        observ = r["observ"]

        ws.cell(row=row, column=1).value = num
        if isinstance(fecha_val, date):
            ws.cell(row=row, column=2).value = fecha_val.strftime("%d/%m/%Y")
        ws.cell(row=row, column=3).value = hora_desde.strftime("%H:%M") if hora_desde else ""
        ws.cell(row=row, column=4).value = hora_hasta.strftime("%H:%M") if hora_hasta else ""
        ws.cell(row=row, column=5).value = tipo_op
        ws.cell(row=row, column=6).value = lugar
        ws.cell(row=row, column=7).value = pers_ident
        ws.cell(row=row, column=8).value = pers_asist
        ws.cell(row=row, column=9).value = delito_prop
        ws.cell(row=row, column=10).value = delito_pers
        ws.cell(row=row, column=11).value = delito_otro
        ws.cell(row=row, column=12).value = lesionados
        ws.cell(row=row, column=13).value = dem_aa
        ws.cell(row=row, column=14).value = dem_av_hecho
        ws.cell(row=row, column=15).value = dem_infraganti
        ws.cell(row=row, column=16).value = dem_contrav
        ws.cell(row=row, column=17).value = rec_hum
        ws.cell(row=row, column=18).value = rec_mat
        ws.cell(row=row, column=19).value = observ

    try:
        wb.save(ruta_excel)
    except PermissionError:
        st.error(
            f"No se pudo guardar el archivo '{ruta_excel}'. "
            "Verifique que no esté abierto en otra aplicación e intente nuevamente."
        )
        st.stop()

    return len(resultados)


def _guardar_por_unidad_anexo2(resultados: List[Dict]) -> int:
    """
    Agrupa los resultados por unidad y los guarda en el archivo
    correspondiente a esa unidad.
    """
    if not resultados:
        return 0

    por_unidad: Dict[str, List[Dict]] = {}
    for r in resultados:
        unidad = r["unidad"].strip()
        por_unidad.setdefault(unidad, []).append(r)

    total = 0
    for unidad, lista in por_unidad.items():
        ruta = UNIT_FILE_ANEXO2.get(unidad)
        if not ruta:
            st.error(
                f"No se encontró un archivo de Excel configurado para la unidad '{unidad}'. "
                "No se guardó ningún dato."
            )
            return 0
        total += _guardar_todos_anexo2(ruta, lista)

    return total


# -------------------------------------------------------------------
# Helpers
# -------------------------------------------------------------------
def _resultado_vacio(vals: Dict) -> bool:
    """
    Consideramos un resultado "vacío" si:
    - tipo_op, lugar y observ están vacíos
    - todos los números en 0
    - horas = 00:00
    """
    if vals["tipo_op"].strip():
        return False
    if vals["lugar"].strip():
        return False
    if vals["observ"].strip():
        return False

    if any(
        vals[campo] != 0
        for campo in [
            "pers_ident",
            "pers_asist",
            "delito_prop",
            "delito_pers",
            "delito_otro",
            "lesionados",
            "dem_aa",
            "dem_av_hecho",
            "dem_infraganti",
            "dem_contrav",
            "rec_hum",
            "rec_mat",
        ]
    ):
        return False

    if vals["hora_desde"] != time(0, 0):
        return False
    if vals["hora_hasta"] != time(0, 0):
        return False

    return True


def _leer_resultado_desde_state(i: int, fecha_carga: date) -> Dict:
    """
    Lee los valores del resultado i desde st.session_state.
    La fecha del operativo es siempre fecha_carga
    (un día antes de la fecha_objetivo).
    """
    unidad = st.session_state.get(f"res_unidad_{i}", "")
    fecha = fecha_carga
    hora_desde = st.session_state.get(f"res_hora_desde_{i}", time(0, 0))
    hora_hasta = st.session_state.get(f"res_hora_hasta_{i}", time(0, 0))
    tipo_op = st.session_state.get(f"res_tipo_op_{i}", "")
    lugar = st.session_state.get(f"res_lugar_{i}", "")
    pers_ident = st.session_state.get(f"res_pers_ident_{i}", 0)
    pers_asist = st.session_state.get(f"res_pers_asist_{i}", 0)
    delito_prop = st.session_state.get(f"res_delito_prop_{i}", 0)
    delito_pers = st.session_state.get(f"res_delito_pers_{i}", 0)
    delito_otro = st.session_state.get(f"res_delito_otro_{i}", 0)
    lesionados = st.session_state.get(f"res_lesionados_{i}", 0)
    dem_aa = st.session_state.get(f"res_dem_aa_{i}", 0)
    dem_av_hecho = st.session_state.get(f"res_dem_av_hecho_{i}", 0)
    dem_infraganti = st.session_state.get(f"res_dem_infraganti_{i}", 0)
    dem_contrav = st.session_state.get(f"res_dem_contrav_{i}", 0)
    rec_hum = st.session_state.get(f"res_rec_hum_{i}", 0)
    rec_mat = st.session_state.get(f"res_rec_mat_{i}", 0)
    observ = st.session_state.get(f"res_observ_{i}", "")

    return {
        "unidad": unidad,
        "fecha": fecha,
        "hora_desde": hora_desde,
        "hora_hasta": hora_hasta,
        "tipo_op": tipo_op,
        "lugar": lugar,
        "pers_ident": pers_ident,
        "pers_asist": pers_asist,
        "delito_prop": delito_prop,
        "delito_pers": delito_pers,
        "delito_otro": delito_otro,
        "lesionados": lesionados,
        "dem_aa": dem_aa,
        "dem_av_hecho": dem_av_hecho,
        "dem_infraganti": dem_infraganti,
        "dem_contrav": dem_contrav,
        "rec_hum": rec_hum,
        "rec_mat": rec_mat,
        "observ": observ,
    }


def _eliminar_resultado(idx: int, cant: int):
    """
    Elimina el resultado idx (0-based) y recorre hacia arriba
    los valores del resto para que no queden "huecos".
    """
    campos = [
        "res_unidad",
        "res_hora_desde",
        "res_hora_hasta",
        "res_tipo_op",
        "res_lugar",
        "res_pers_ident",
        "res_pers_asist",
        "res_delito_prop",
        "res_delito_pers",
        "res_delito_otro",
        "res_lesionados",
        "res_dem_aa",
        "res_dem_av_hecho",
        "res_dem_infraganti",
        "res_dem_contrav",
        "res_rec_hum",
        "res_rec_mat",
        "res_observ",
    ]

    for j in range(idx, cant - 1):
        for c in campos:
            src = f"{c}_{j+1}"
            dst = f"{c}_{j}"
            st.session_state[dst] = st.session_state.get(src)

    last = cant - 1
    for c in campos:
        key = f"{c}_{last}"
        if key in st.session_state:
            del st.session_state[key]

    st.session_state["anexo2_cant_resultados"] = cant - 1
    st.rerun()


# -------------------------------------------------------------------
# Pantalla principal ANEXO II
# -------------------------------------------------------------------
def mostrar_anexo_2(ruta_excel: str, fecha_objetivo: date, fecha_default: date) -> bool:
    """
    ruta_excel se ignora ahora (se usan archivos por unidad).
    """

    unidad_actual = st.session_state.get("unidad_actual", "")

    # Fecha de los operativos que se cargan (día anterior a la fecha_objetivo)
    fecha_carga = fecha_objetivo - timedelta(days=1)

    if "anexo2_cant_resultados" not in st.session_state:
        st.session_state["anexo2_cant_resultados"] = 1

    cant = st.session_state["anexo2_cant_resultados"]

    st.title("BIENVENIDO AL SISTEMA DE CARGA OPERATIVOS VERANO")
    st.markdown("### RESULTADOS DEL OPERATIVO")
    st.write(
        f"Carga de resultados correspondiente al día **{fecha_carga.strftime('%d/%m/%Y')}** "
        f"(día anterior a la fecha objetivo {fecha_objetivo.strftime('%d/%m/%Y')})."
    )

    st.caption(
        "Complete los campos de cada resultado. Puede agregar más resultados con el botón "
        "**Agregar otro resultado**. Los campos numéricos pueden ser 0, excepto "
        "**Recurso humano** y **Recurso material**, que deben ser mayores a 0. "
        "Las horas deben tener minutos en 00. La fecha del operativo se fija "
        "automáticamente (día anterior a la fecha objetivo) y no se puede modificar. "
        "Los datos se guardan en el archivo correspondiente a la unidad seleccionada."
    )

    for i in range(cant):
        st.markdown(f"#### Resultado {i + 1}")
        with st.container(border=True):

            st.session_state[f"res_unidad_{i}"] = unidad_actual
            st.text_input(
                "Unidad",
                value=unidad_actual,
                key=f"res_unidad_{i}",
                disabled=True,
            )

            st.date_input(
                "Fecha del operativo",
                value=fecha_carga,
                key=f"res_fecha_{i}",
                disabled=True,
            )

            col_h1, col_h2 = st.columns(2)
            with col_h1:
                st.time_input(
                    "Hora desde",
                    value=time(0, 0),
                    step=timedelta(hours=1),
                    key=f"res_hora_desde_{i}",
                )
            with col_h2:
                st.time_input(
                    "Hora hasta",
                    value=time(0, 0),
                    step=timedelta(hours=1),
                    key=f"res_hora_hasta_{i}",
                )

            st.text_input("Tipo de operativo", key=f"res_tipo_op_{i}")
            st.text_input("Lugar", key=f"res_lugar_{i}")

            col1, col2 = st.columns(2)
            with col1:
                st.number_input(
                    "Personas identificadas",
                    min_value=0,
                    step=1,
                    value=0,
                    key=f"res_pers_ident_{i}",
                )
            with col2:
                st.number_input(
                    "Personas asistidas",
                    min_value=0,
                    step=1,
                    value=0,
                    key=f"res_pers_asist_{i}",
                )

            st.markdown("**Modalidad de intervención**")
            col_m1, col_m2, col_m3 = st.columns(3)
            with col_m1:
                st.number_input(
                    "Delito contra la propiedad",
                    min_value=0,
                    step=1,
                    value=0,
                    key=f"res_delito_prop_{i}",
                )
            with col_m2:
                st.number_input(
                    "Delito contra las personas",
                    min_value=0,
                    step=1,
                    value=0,
                    key=f"res_delito_pers_{i}",
                )
            with col_m3:
                st.number_input(
                    "Otro",
                    min_value=0,
                    step=1,
                    value=0,
                    key=f"res_delito_otro_{i}",
                )

            st.number_input(
                "Lesionados",
                min_value=0,
                step=1,
                value=0,
                key=f"res_lesionados_{i}",
            )

            st.markdown("**Demorados**")
            col_d1, col_d2, col_d3, col_d4 = st.columns(4)
            with col_d1:
                st.number_input(
                    "AA",
                    min_value=0,
                    step=1,
                    value=0,
                    key=f"res_dem_aa_{i}",
                )
            with col_d2:
                st.number_input(
                    "AV. de hecho",
                    min_value=0,
                    step=1,
                    value=0,
                    key=f"res_dem_av_hecho_{i}",
                )
            with col_d3:
                st.number_input(
                    "Infraganti",
                    min_value=0,
                    step=1,
                    value=0,
                    key=f"res_dem_infraganti_{i}",
                )
            with col_d4:
                st.number_input(
                    "Contravención",
                    min_value=0,
                    step=1,
                    value=0,
                    key=f"res_dem_contrav_{i}",
                )

            st.markdown("**Recursos**")
            col_r1, col_r2 = st.columns(2)
            with col_r1:
                st.number_input(
                    "Recurso humano",
                    min_value=0,
                    step=1,
                    value=0,
                    key=f"res_rec_hum_{i}",
                )
            with col_r2:
                st.number_input(
                    "Recurso material",
                    min_value=0,
                    step=1,
                    value=0,
                    key=f"res_rec_mat_{i}",
                )

            st.text_area(
                "Minuta del hecho / Observaciones",
                key=f"res_observ_{i}",
            )

            if cant > 1:
                if st.button("Eliminar este resultado", key=f"del_res_{i}"):
                    _eliminar_resultado(i, cant)

    st.markdown("---")
    col_add, col_fin = st.columns(2)

    if col_add.button("Agregar otro resultado"):
        st.session_state["anexo2_cant_resultados"] = cant + 1
        st.rerun()

    finalizado = False

    if col_fin.button("Guardar y FINALIZAR LA CARGA"):
        errores: List[str] = []
        resultados_a_guardar: List[Dict] = []
        resultados_vacios: List[int] = []

        for i in range(cant):
            vals = _leer_resultado_desde_state(i, fecha_carga)

            if _resultado_vacio(vals):
                resultados_vacios.append(i + 1)
                continue

            if not vals["unidad"].strip():
                errores.append(f"Resultado {i+1}: Unidad es obligatoria.")
            if vals["unidad"].strip() not in UNIT_FILE_ANEXO2:
                errores.append(
                    f"Resultado {i+1}: Unidad '{vals['unidad']}' no tiene archivo configurado."
                )

            if not vals["tipo_op"].strip():
                errores.append(f"Resultado {i+1}: Tipo de operativo es obligatorio.")
            if not vals["lugar"].strip():
                errores.append(f"Resultado {i+1}: Lugar es obligatorio.")

            if vals["rec_hum"] <= 0:
                errores.append(
                    f"Resultado {i+1}: Recurso humano debe ser mayor a 0."
                )
            if vals["rec_mat"] <= 0:
                errores.append(
                    f"Resultado {i+1}: Recurso material debe ser mayor a 0."
                )

            hd = vals["hora_desde"]
            hh = vals["hora_hasta"]
            if hd.minute != 0 or hh.minute != 0:
                errores.append(
                    f"Resultado {i+1}: las horas deben ser en números "
                    "redondos (minuto 00)."
                )
            if hd == time(0, 0) or hh == time(0, 0):
                errores.append(
                    f"Resultado {i+1}: las horas no pueden ser 00:00."
                )

            resultados_a_guardar.append(vals)

        if resultados_vacios:
            lista = ", ".join(str(n) for n in resultados_vacios)
            errores.append(
                f"Los resultados {lista} están vacíos. "
                "Elimínelos o complételos antes de finalizar."
            )

        if not resultados_a_guardar and not errores:
            errores.append(
                "No hay ningún resultado completo para guardar. "
                "Complete al menos un resultado."
            )

        if errores:
            st.error("Se encontraron errores en la carga:")
            for e in errores:
                st.write(f"- {e}")
        else:
            n = _guardar_por_unidad_anexo2(resultados_a_guardar)
            if n == 0:
                st.warning(
                    "No se guardó ningún resultado. "
                    "Revise los datos e intente nuevamente."
                )
            else:
                if n == 1:
                    msg = (
                        "Se ha guardado 1 resultado del día y "
                        "se finalizó la carga de resultados."
                    )
                else:
                    msg = (
                        f"Se han guardado {n} resultados del día y "
                        "se finalizó la carga de resultados."
                    )

                st.session_state["anexo2_flash_msg"] = msg
                finalizado = True

    return finalizado