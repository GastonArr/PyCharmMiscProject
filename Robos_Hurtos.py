# Robos_Hurtos.py
# Subflujo para delitos de Robos / Hurtos
# Inserta pantallas adicionales entre Direcciones (step=5) y Resumen (step=6) del app principal.
# Escribe en la fila activa (param "fila") del Excel abierto en "excel_path".

import streamlit as st
from openpyxl import load_workbook, Workbook
import os

# ==============================
# Utilidades de Excel
# ==============================
def is_xlsm(path: str) -> bool:
    return path.lower().endswith(".xlsm")

def asegurar_excel(path: str):
    carpeta = os.path.dirname(path)
    if carpeta and not os.path.exists(carpeta):
        os.makedirs(carpeta, exist_ok=True)
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Hoja1"
        wb.save(path)

def cargar_libro(path: str):
    asegurar_excel(path)
    return load_workbook(path, keep_vba=is_xlsm(path))

def unwrap_quotes(v):
    """Quita SOLO comillas envolventes si existen; no toca espacios internos."""
    if not isinstance(v, str):
        return v
    if len(v) >= 2 and ((v[0] == v[-1] == '"') or (v[0] == v[-1] == "'")):
        return v[1:-1]
    return v

def _trim(s):
    return s.strip() if isinstance(s, str) else s

def mostrar_hecho_referencia():
    """Muestra el hecho cargado en el flujo principal para tenerlo siempre a la vista."""
    hecho = st.session_state.get("hecho")
    if hecho:
        st.subheader("Hecho ingresado para referencia:")
        st.write(hecho)
        st.markdown("---")

# ==============================
# Delitos que activan este flujo
# ==============================
DELITOS_APLICA = {
    "ROBO SIMPLE ",
    "HURTO SIMPLE ",
    "ROBO AGRAVADO POR EL USO DE ARMA DE ARMA BLANCA",
    "ROBO AGRAVADO POR EL USO DE ARMA DE FUEGO",
    "ROBO ABIGEATO ",
    "ROBO AGRAVADO POR LESION ",
}
DELITOS_APLICA_STRIPPED = {d.strip() for d in DELITOS_APLICA}

# ==============================
# Catálogos
# ==============================
SEXO_OPCIONES = ["MASCULINO", "FEMENINO", "NO CONSTA"]
CANT_1_10 = [str(i) for i in range(1, 11)]

VULNERABILIDAD = [
    "Adultos Mayores",
    "Niñez - Edad joven",
    "Mujer embarazada",
    "Persona con discapacidad",
    "Personas en situacion de calle",
    "Migrantes",
    "Minorías Etnicas, Raciales o Religiosas",
    "No pertenece a ningun grupo vulnerable",
    "Sin determinar",
]

TIPO_ARMA = [
    "Hechos con ARMA DE FUEGO*",
    "Hechos con OTRA ARMA*",
    "Hechos SIN ARMAS*",
    "SIN DETERMINAR *",
]

INCULPADOS_SN = ["NO", "SI"]  # NO primero para que quede preseleccionado
INCULPADOS_RANGO = ["Hasta 15 año", "15 a 17 años", "mayor de 18 años", "Sin Determinar"]

ELEMENTOS_SUSTRAIDOS = ["AUTOMOTOR", "MOTOCICLETA", "BANCO", "ABIGEATO", "OTROS"]

TIPO_LUGAR_BASE = [
    "HECHOS EN VÍA PÚBLICA (CALLE, PLAZA,ETC)",
    "HECHOS EN ESTABLECIMIENTOS COMERCIALES/PÚBLICOS",
    "HECHOS EN DOMICILIO PARTICULAR",
    "EN TRANSPORTE (SOLO ABIGEATO)",
    "SIN DETERMINAR",
]

DETALLE_EST_PUBLICOS = [
    "ESCUELA",
    "SHOPPING",
    "UNIVERSIDAD",
    "MERCADO/ DESPENSA",
    "OFICINA MUNICIPAL",
    "OFICINA PROVINCIAL",
    "OFICINA NACIONAL",
    "KIOSCO/ MAXIKIOSCO",
    "HOSPITAL",
    "ESTABLECIMIENTO DEPORTIVO",
    "DEPENDENCIA POLICIAL",
    "BASE PETROLERA",
    "POLIDEPORTIVO",
    "YACIMIENTO",
    "BODEGA",
    "OTROS (Detallar en Comentarios)",
]
DETALLE_EST_VIA_PUBLICA = [
    "CALLE",
    "PARQUE",
    "PLAZA",
    "ESTACIONAMIENTO",
    "PARADA DE COLECTIVO",
    "COSTANERA",
    "DESCAMPADO",
    "OTROS (Detallar en Comentarios)",
    "CHACRA",
]

SUBCAT_POR_ELEMENTO = {
    "AUTOMOTOR": [
        "FIAT","CHEVROLET","MERCEDES_BENZ","NISSAN","PEUGEOT","FORD","IVECO","CITROEN",
        "RENAULT","SCANIA","TOYOTA","VOLKSWAGEN","CHERY","DODGE","JEEP","HYUNDAI","OTRAS_MARCAS",
    ],
    "MOTOCICLETA": ["MOTO","TRICICLO","CUATRICICLO","OTRO (Detallar en Comentarios)"],
    "ABIGEATO": ["BOVINO","AVIAR","EQUINO","CAPRINO","PORCINO","OVINO","PISCICOLA","EXOTICO"],
    "BANCO": [],
    "OTROS": [
        "BILLETERA","CELULAR","BICICLETA","NOTEBOOK","ELECTRODOMESTICOS (Detallar en comentarios)",
        "RUEDA DE AUTO","BATERIA DE AUTO","DINERO_NACIONAL","DINERO_EXTRANJERO (Detallar en comentarios)",
        "MOCHILA - CARTERA","JOYAS","HERRAMIENTAS",
        "METAL (Cobre, Aluminio, Etc, Detallar en  Comentarios)",
        "OTROS (Detallar en comentarios)",
    ],
}

DENOMINACION_POR_SUBCAT = {
    "FIAT": ["COUPE","SEDAN 2/3 PUERTAS","SEDAN 4/5 PUERTAS","PICK UP 2/3 PUERTAS","PICK UP 4/5 PUERTAS",
             "DESCAPOTABLE O CONVERTIBLE","MONOVOLUMEN","UTILITARIO"],
    "CHEVROLET": ["COUPE","SEDAN 2/3 PUERTAS","SEDAN 4/5 PUERTAS","PICK UP 2/3 PUERTAS","PICK UP 4/5 PUERTAS",
                  "DESCAPOTABLE O CONVERTIBLE","MONOVOLUMEN","UTILITARIO"],
    "MERCEDES_BENZ": ["COUPE","SEDAN 2/3 PUERTAS","SEDAN 4/5 PUERTAS","PICK UP 2/3 PUERTAS","PICK UP 4/5 PUERTAS",
                      "DESCAPOTABLE O CONVERTIBLE","MONOVOLUMEN","UTILITARIO","TRACTOR(CAMION)","MAQUINARIA AGRICOLA"],
    "NISSAN": ["COUPE","SEDAN 2/3 PUERTAS","SEDAN 4/5 PUERTAS","PICK UP 2/3 PUERTAS","PICK UP 4/5 PUERTAS",
               "DESCAPOTABLE O CONVERTIBLE","MONOVOLUMEN","UTILITARIO"],
    "PEUGEOT": ["COUPE","SEDAN 2/3 PUERTAS","SEDAN 4/5 PUERTAS","PICK UP 2/3 PUERTAS","PICK UP 4/5 PUERTAS",
                "DESCAPOTABLE O CONVERTIBLE","MONOVOLUMEN","UTILITARIO"],
    "FORD": ["COUPE","SEDAN 2/3 PUERTAS","SEDAN 4/5 PUERTAS","PICK UP 2/3 PUERTAS","PICK UP 4/5 PUERTAS",
             "DESCAPOTABLE O CONVERTIBLE","MONOVOLUMEN","UTILITARIO","TRACTOR(CAMION)","MAQUINARIA AGRICOLA"],
    "IVECO": ["UTILITARIO","TRACTOR(CAMION)","MAQUINARIA AGRICOLA"],
    "CITROEN": ["COUPE","SEDAN 2/3 PUERTAS","SEDAN 4/5 PUERTAS","PICK UP 2/3 PUERTAS","PICK UP 4/5 PUERTAS",
                "DESCAPOTABLE O CONVERTIBLE","MONOVOLUMEN","UTILITARIO"],
    "RENAULT": ["COUPE","SEDAN 2/3 PUERTAS","SEDAN 4/5 PUERTAS","PICK UP 2/3 PUERTAS","PICK UP 4/5 PUERTAS",
                "DESCAPOTABLE O CONVERTIBLE","MONOVOLUMEN","UTILITARIO"],
    "SCANIA": ["UTILITARIO","TRACTOR(CAMION)","MAQUINARIA AGRICOLA"],
    "TOYOTA": ["COUPE","SEDAN 2/3 PUERTAS","SEDAN 4/5 PUERTAS","PICK UP 2/3 PUERTAS","PICK UP 4/5 PUERTAS",
               "DESCAPOTABLE O CONVERTIBLE","MONOVOLUMEN","UTILITARIO","TRACTOR(CAMION)","MAQUINARIA AGRICOLA"],
    "VOLKSWAGEN": ["COUPE","SEDAN 2/3 PUERTAS","SEDAN 4/5 PUERTAS","PICK UP 2/3 PUERTAS","PICK UP 4/5 PUERTAS",
                   "DESCAPOTABLE O CONVERTIBLE","MONOVOLUMEN","UTILITARIO"],
    "CHERY": ["COUPE","SEDAN 2/3 PUERTAS","SEDAN 4/5 PUERTAS","PICK UP 2/3 PUERTAS","PICK UP 4/5 PUERTAS",
              "DESCAPOTABLE O CONVERTIBLE","MONOVOLUMEN","UTILITARIO"],
    "DODGE": ["COUPE","SEDAN 2/3 PUERTAS","SEDAN 4/5 PUERTAS","PICK UP 2/3 PUERTAS","PICK UP 4/5 PUERTAS",
              "DESCAPOTABLE O CONVERTIBLE","MONOVOLUMEN","UTILITARIO"],
    "JEEP": ["COUPE","SEDAN 2/3 PUERTAS","SEDAN 4/5 PUERTAS","PICK UP 2/3 PUERTAS","PICK UP 4/5 PUERTAS",
             "DESCAPOTABLE O CONVERTIBLE","MONOVOLUMEN","UTILITARIO"],
    "HYUNDAI": ["COUPE","SEDAN 2/3 PUERTAS","SEDAN 4/5 PUERTAS","PICK UP 2/3 PUERTAS","PICK UP 4/5 PUERTAS",
                "DESCAPOTABLE O CONVERTIBLE","MONOVOLUMEN","UTILITARIO"],
    "OTRAS_MARCAS": ["COUPE","SEDAN 2/3 PUERTAS","SEDAN 4/5 PUERTAS","PICK UP 2/3 PUERTAS","PICK UP 4/5 PUERTAS",
                     "DESCAPOTABLE O CONVERTIBLE","MONOVOLUMEN","UTILITARIO"],
}

MODUS_HURTO = [
    "OPORTUNISTA","MECHERAS","PUNGUISTA - LANCERO","DESMANTELADORES",
    "EMPLEADO INFIEL","MOTOCHORRO","PIRAÑA","INHIBIDORES",
]
MODUS_ROBO = [
    "ESCRUCHANTE_Violacion_de_domicilio","ASALTANTE","DESMANTELADORES",
    "ESPIANTADORES_Sustraccion_de_vehiculo","CUATREROS_Abigeato",
    "MOTOCHORRO","PIRAÑA","OPORTUNISTA",
]
ESPECIALIDAD_POR_MODUS = {
    "ESCRUCHANTE_Violacion_de_domicilio": [
        "Violentador de CERROJOS Y PUERTAS",
        "Violentador de VENTANAS, VENTILUCES Y TECHOS",
        "Violentador de VIDRIERAS",
        "Violentador de PUERTAS Y VENTANAS DE AUTOMOTORES",
        "ENTRADERAS",
        "BOQUETEROS",
    ],
    "ASALTANTE": [
        "Violentador de CERROJOS Y PUERTAS",
        "ATRACADORES",
        "Violentador de CAJAS FUERTES",
        "SALIDERA DE BANCOS",
    ],
    "ESPIANTADORES_Sustraccion_de_vehiculo": ["PIRATAS DEL ASFALTO"],
    "CUATREROS_Abigeato": [],
    "MOTOCHORRO": [],
    "PIRAÑA": [],
    "OPORTUNISTA": [],
    "DESMANTELADORES": [],
}

# ==============================
# Render principal
# ==============================
def render(excel_path: str, fila: int, delito_x3: str) -> None:
    """
    Inserta pantallas adicionales para Robos/Hurtos entre Direcciones y Resumen.
    Maneja su propio subflujo con st.session_state.rh_step.
    """
    delito_norm = (delito_x3 or "").strip()
    if delito_norm not in DELITOS_APLICA_STRIPPED:
        st.session_state.step = 6
        st.rerun()

    if "rh_step" not in st.session_state:
        st.session_state.rh_step = 1

    # === Precarga desde rh_preview/rh_cache para conservar datos al volver ===
    ss = st.session_state
    prev = ss.get("rh_preview") or {}
    cache = ss.get("rh_cache") or {}

    # Si no hay cache pero sí hay preview (volvés desde el Resumen), reconstruimos cache
    if not cache and prev:
        ss.rh_cache = {
            "vict_rows": list(prev.get("vict_rows") or []),
            "vict_total": prev.get("vict_total"),
            "vulnerab": prev.get("vulnerab"),
            "tipo_arma": prev.get("tipo_arma"),
            "inc_sn": prev.get("inc_sn"),
            "rango_etario": prev.get("rango_etario"),
            "cant_rango": prev.get("cant_rango"),
            "sex_rows": list(prev.get("sex_rows") or []),
            "tipo_lugar": prev.get("tipo_lugar"),
            "detalle_est": prev.get("detalle_est"),
            "elem": prev.get("elem"),
            "subcat": prev.get("subcat"),
            "denom": prev.get("denom"),
            "anio": prev.get("anio"),
            "modelo": prev.get("modelo"),
        }
        cache = ss.rh_cache

    # Sembrar listas dinámicas (víctimas / inculpados) si no están presentes
    if ("rh_vict_rows" not in ss or not isinstance(ss.rh_vict_rows, list)) and (cache.get("vict_rows")):
        ss.rh_vict_rows = list(cache.get("vict_rows"))
    if ("rh_sex_rows" not in ss or not isinstance(ss.rh_sex_rows, list)) and (cache.get("sex_rows")):
        ss.rh_sex_rows = list(cache.get("sex_rows"))

    # Helper para no pisar lo que el usuario acaba de escribir
    def seed(k, v):
        if v not in (None, "") and k not in ss:
            ss[k] = v

    # Sembrar claves simples usadas por los widgets (Paso 1)
    seed("rh_vulnerabilidad", cache.get("vulnerab"))
    # Si NO es Hurto Simple, restauramos tipo de arma; para Hurto se forza SIN ARMAS más adelante
    if (ss.get("delito") or delito_x3 or "").strip() != "HURTO SIMPLE":
        seed("rh_tipo_arma", cache.get("tipo_arma"))
    seed("rh_inc_sn", cache.get("inc_sn") or "NO")
    seed("rh_inc_rango", cache.get("rango_etario"))
    seed("rh_inc_cant_rango", cache.get("cant_rango"))
    seed("rh_tipo_lugar", cache.get("tipo_lugar"))
    seed("rh_detalle_est", cache.get("detalle_est"))
    seed("rh_elemento", cache.get("elem"))
    seed("rh_subcat", cache.get("subcat"))
    seed("rh_denom", cache.get("denom"))
    seed("rh_anio", cache.get("anio"))
    seed("rh_modelo", cache.get("modelo"))
    # Paso 2
    seed("rh_modus", prev.get("modus"))
    seed("rh_especialidad", prev.get("especialidad"))

    def C(col: str) -> str:
        return f"{col}{fila}"

    # =================== Paso 1: Datos generales ===================
    if st.session_state.rh_step == 1:
        st.subheader("Datos adicionales (Robos/Hurtos)")
        mostrar_hecho_referencia()

        # ---------- Víctimas: distribución por sexo (dinámica) ----------
        st.markdown("### Víctimas")
        st.caption("Distribución por sexo de las víctimas. Podés agregar otro sexo si corresponde (máx. 3 filas).")

        if "rh_vict_rows" not in st.session_state or not isinstance(st.session_state.rh_vict_rows, list):
            st.session_state.rh_vict_rows = [{"sexo": "MASCULINO", "cant": "1"}]  # renglón inicial

        total_vict = 0
        vict_borrar = []
        for i, row in enumerate(st.session_state.rh_vict_rows):
            c1, c2, c3 = st.columns([2, 1, 1])
            with c1:
                vict_sexo = st.selectbox(
                    "Sexo",
                    SEXO_OPCIONES,
                    index=SEXO_OPCIONES.index(row.get("sexo", "MASCULINO")) if row.get("sexo") in SEXO_OPCIONES else 0,
                    key=f"rh_vict_rows_{i}_sexo",
                )
            with c2:
                vict_cant = st.selectbox(
                    "Cantidad",
                    CANT_1_10,
                    index=(CANT_1_10.index(row.get("cant")) if row.get("cant") in CANT_1_10 else 0),
                    key=f"rh_vict_rows_{i}_cant",
                )
            with c3:
                if i > 0:
                    if st.button("Quitar", key=f"rh_vict_rows_{i}_del"):
                        vict_borrar.append(i)
            st.session_state.rh_vict_rows[i]["sexo"] = vict_sexo
            st.session_state.rh_vict_rows[i]["cant"] = vict_cant
            total_vict += int(vict_cant)

        if vict_borrar:
            st.session_state.rh_vict_rows = [r for j, r in enumerate(st.session_state.rh_vict_rows) if j not in vict_borrar]
            st.rerun()

        if len(st.session_state.rh_vict_rows) < 3 and total_vict < 10:
            if st.button("Agregar otro sexo", key="rh_add_vict_row"):
                usados = {r["sexo"] for r in st.session_state.rh_vict_rows if "sexo" in r}
                restantes = [s for s in SEXO_OPCIONES if s not in usados]
                nuevo_sexo = restantes[0] if restantes else SEXO_OPCIONES[0]
                st.session_state.rh_vict_rows.append({"sexo": nuevo_sexo, "cant": "1"})
                st.rerun()

        st.caption(f"Total víctimas (AO): {total_vict}")

        # --- Vulnerabilidad
        vulnerab = st.selectbox("VULNERABILIDAD DE LA VICTIMAS", VULNERABILIDAD, key="rh_vulnerabilidad")

        # --- Tipo de arma (HURTO SIMPLE => SIN ARMAS)
        if delito_norm == "HURTO SIMPLE":
            tipo_arma = "Hechos SIN ARMAS*"
            st.info("Delito HURTO SIMPLE: se establece automáticamente 'Hechos SIN ARMAS*' (AR).")
        else:
            tipo_arma = st.selectbox("TIPO DE ARMA", TIPO_ARMA, key="rh_tipo_arma")

        st.markdown("---")

        # ---------- Inculpados ----------
        st.markdown("### Inculpados")
        inc_sn = st.radio(
            "Indique si hay personas que son acusadas o sospechosas de haber cometido un delito (por policía o autoridad judicial), "
            "es decir, quienes enfrentan imputaciones legales durante un procedimiento penal.",
            INCULPADOS_SN,
            horizontal=True,
            index=0,
            key="rh_inc_sn",
        )

        rango_etario = None
        cant_rango = None

        if inc_sn == "SI":
            colr, colc = st.columns([2, 1])
            with colr:
                rango_etario = st.selectbox(
                    "Hasta 15 año / 15 a 17 años / mayor de 18 años / Sin Determinar",
                    INCULPADOS_RANGO,
                    key="rh_inc_rango",
                )
            with colc:
                cant_rango = st.selectbox(
                    "indique el número de personas según la edad seleccionada",
                    CANT_1_10,
                    key="rh_inc_cant_rango",
                )

            st.markdown("**Distribución por sexo de inculpados**")

            if "rh_sex_rows" not in st.session_state or not isinstance(st.session_state.rh_sex_rows, list):
                st.session_state.rh_sex_rows = [{"sexo": "MASCULINO", "cant": "1"}]

            total_inc_sex = 0
            inc_borrar = []
            for i, row in enumerate(st.session_state.rh_sex_rows):
                c1, c2, c3 = st.columns([2, 1, 1])
                with c1:
                    sexo_val = st.selectbox(
                        "Sexo",
                        SEXO_OPCIONES,
                        index=SEXO_OPCIONES.index(row.get("sexo", "MASCULINO")) if row.get("sexo") in SEXO_OPCIONES else 0,
                        key=f"rh_sex_rows_{i}_sexo",
                    )
                with c2:
                    cant_val = st.selectbox(
                        "ingrese la cantidad según el sexo",
                        CANT_1_10,
                        index=(CANT_1_10.index(row.get("cant")) if row.get("cant") in CANT_1_10 else 0),
                        key=f"rh_sex_rows_{i}_cant",
                    )
                with c3:
                    if i > 0:
                        if st.button("Quitar", key=f"rh_sex_rows_{i}_del"):
                            inc_borrar.append(i)
                st.session_state.rh_sex_rows[i]["sexo"] = sexo_val
                st.session_state.rh_sex_rows[i]["cant"] = cant_val
                total_inc_sex += int(cant_val)

            if inc_borrar:
                st.session_state.rh_sex_rows = [r for j, r in enumerate(st.session_state.rh_sex_rows) if j not in inc_borrar]
                st.rerun()

            try:
                objetivo_inc = int(cant_rango) if cant_rango else 0
            except Exception:
                objetivo_inc = 0

            if len(st.session_state.rh_sex_rows) < 3 and total_inc_sex < objetivo_inc:
                if st.button("Agregar otro sexo", key="rh_add_sex_row"):
                    usados = {r["sexo"] for r in st.session_state.rh_sex_rows if "sexo" in r}
                    restantes = [s for s in SEXO_OPCIONES if s not in usados]
                    nuevo_sexo = restantes[0] if restantes else SEXO_OPCIONES[0]
                    st.session_state.rh_sex_rows.append({"sexo": nuevo_sexo, "cant": "1"})
                    st.rerun()

            st.caption(f"Total asignado por sexo: {total_inc_sex} / {objetivo_inc}. Restante: {max(0, objetivo_inc - total_inc_sex)}.")
        else:
            st.session_state.rh_sex_rows = []

        st.markdown("---")

        # --- TIPO DE LUGAR (Hurto Simple no habilita transporte)
        opciones_tipo_lugar = list(TIPO_LUGAR_BASE)
        if delito_norm == "HURTO SIMPLE":
            opciones_tipo_lugar = [x for x in opciones_tipo_lugar if x != "EN TRANSPORTE (SOLO ABIGEATO)"]
        tipo_lugar = st.selectbox("TIPO DE LUGAR", opciones_tipo_lugar, key="rh_tipo_lugar")

        # Detalle establecimiento (solo si aplica)
        detalle_est = None
        if tipo_lugar == "HECHOS EN ESTABLECIMIENTOS COMERCIALES/PÚBLICOS":
            detalle_est = st.selectbox("DETALLE ESTABLECIMIENTO", DETALLE_EST_PUBLICOS, key="rh_detalle_est")
        elif tipo_lugar == "HECHOS EN VÍA PÚBLICA (CALLE, PLAZA,ETC)":
            detalle_est = st.selectbox("DETALLE ESTABLECIMIENTO", DETALLE_EST_VIA_PUBLICA, key="rh_detalle_est")
        else:
            st.caption("DETALLE ESTABLECIMIENTO: (no aplica)")

        st.markdown("---")

        # --- ELEMENTOS SUSTRAIDOS (SIEMPRE SE MUESTRA)
        elem = st.selectbox("ELEMENTOS SUSTRAIDOS", ELEMENTOS_SUSTRAIDOS, key="rh_elemento")

        # Subcategoría / Denominación / Año / Modelo
        subcat = None
        denom = None
        anio = None
        modelo = None

        subcat_list = SUBCAT_POR_ELEMENTO.get(elem, [])
        if subcat_list:
            subcat = st.selectbox("SUBCATEGORIA", subcat_list, key="rh_subcat")
            if elem == "AUTOMOTOR" and subcat in DENOMINACION_POR_SUBCAT:
                denom = st.selectbox("DENOMINACION", DENOMINACION_POR_SUBCAT[subcat], key="rh_denom")
            else:
                st.caption("DENOMINACION: (no aplica)")
        else:
            st.caption("SUBCATEGORIA: (no aplica)")

        if elem in ("AUTOMOTOR", "MOTOCICLETA"):
            colA, colB = st.columns(2)
            with colA:
                anio = st.text_input("AÑO (4 dígitos)",
                                     value=st.session_state.get("rh_anio", ""),
                                     max_chars=4, key="rh_anio")
            with colB:
                modelo = st.text_input("MODELO (máx 20)",
                                       value=st.session_state.get("rh_modelo", ""),
                                       max_chars=20, key="rh_modelo")
        else:
            st.caption("AÑO / MODELO: (no aplica)")

        # --- Navegación
        col1, col2 = st.columns(2)
        with col1:
            if st.button("Volver ⬅️", key="rh_volver_1"):
                st.session_state.step = 5
                st.rerun()
        with col2:
            if st.button("Siguiente ➡️", key="rh_siguiente_1"):
                # Validaciones mínimas (tuyas, intactas)

                # Víctimas: total obligatorio entre 1 y 10
                total_vict = sum(int(r.get("cant", "0") or "0") for r in st.session_state.rh_vict_rows)
                if total_vict < 1:
                    st.warning("Debe indicar al menos 1 víctima.")
                    st.stop()
                if total_vict > 10:
                    st.warning("El total de víctimas no puede superar 10.")
                    st.stop()

                # VALIDACIÓN OBLIGATORIA: AÑO y MODELO si Automotor/Motocicleta
                if elem in ("AUTOMOTOR", "MOTOCICLETA"):
                    if not anio or not anio.isdigit() or len(anio) != 4:
                        st.warning("Ingrese el AÑO del vehículo (4 dígitos).")
                        st.stop()
                    if not modelo or (isinstance(modelo, str) and modelo.strip() == ""):
                        st.warning("Ingrese el MODELO del vehículo.")
                        st.stop()

                # Inculpados: si hay, validar igualdad suma por sexo == cantidad por rango
                if inc_sn == "SI":
                    if not (rango_etario and cant_rango):
                        st.warning("Complete el rango etario y la cantidad para inculpados.")
                        st.stop()
                    objetivo_inc = int(cant_rango)
                    total_inc = sum(int(r.get("cant", "0") or "0") for r in st.session_state.rh_sex_rows)
                    if total_inc != objetivo_inc:
                        st.warning("seleccione el mismo numero el cual indico en “indique el número de personas según la edad seleccionada”.")
                        st.stop()

                # Cache y avanzar
                st.session_state.rh_cache = {
                    "vict_rows": list(st.session_state.rh_vict_rows),
                    "vict_total": str(total_vict),
                    "vulnerab": vulnerab,
                    "tipo_arma": tipo_arma,
                    "inc_sn": inc_sn,
                    "rango_etario": rango_etario,
                    "cant_rango": cant_rango,
                    "sex_rows": list(st.session_state.rh_sex_rows) if inc_sn == "SI" else [],
                    "tipo_lugar": tipo_lugar,
                    "detalle_est": detalle_est,
                    "elem": elem,
                    "subcat": subcat,
                    "denom": denom,
                    "anio": anio,
                    "modelo": modelo,
                }
                st.session_state.rh_step = 2
                st.rerun()

    # =================== Paso 2: Modus & Especialidad + Guardar ===================
    elif st.session_state.rh_step == 2:
        # Mostrar hecho de referencia también en este paso
        mostrar_hecho_referencia()

        cache = st.session_state.get("rh_cache", {})

        # Modus según delito: Hurto vs Robo
        delito_norm = (st.session_state.get("delito") or "").strip() or (delito_x3 or "").strip()
        es_hurto = (delito_norm == "HURTO SIMPLE")
        opciones_modus = MODUS_HURTO if es_hurto else MODUS_ROBO
        modus = st.selectbox("MODUS OPERANDI", opciones_modus, key="rh_modus")

        esp_list = ESPECIALIDAD_POR_MODUS.get(modus, [])
        if esp_list:
            especialidad = st.selectbox("ESPECIALIDAD", esp_list, key="rh_especialidad")
        else:
            especialidad = None
            st.caption("ESPECIALIDAD: (no aplica)")

        st.markdown("---")

        col1, col2 = st.columns(2)
        with col1:
            if st.button("Volver ⬅️", key="rh_volver_2"):
                st.session_state.rh_step = 1
                st.rerun()
        with col2:
            if st.button("Guardar y continuar ➡️", key="rh_guardar"):
                try:
                    # ⛔️ SIN guardar en Excel acá. Solo armamos el preview y avanzamos.
                    rh_preview = {
                        "vict_rows": cache.get("vict_rows"),
                        "vict_total": cache.get("vict_total"),
                        "vulnerab": cache.get("vulnerab"),
                        "tipo_arma": cache.get("tipo_arma"),
                        "inc_sn": cache.get("inc_sn"),
                        "rango_etario": cache.get("rango_etario"),
                        "cant_rango": cache.get("cant_rango"),
                        "sex_rows": cache.get("sex_rows"),
                        "tipo_lugar": cache.get("tipo_lugar"),
                        "detalle_est": cache.get("detalle_est"),
                        "elem": cache.get("elem"),
                        "subcat": cache.get("subcat"),
                        "denom": cache.get("denom"),
                        "anio": cache.get("anio"),
                        "modelo": cache.get("modelo"),
                        "modus": modus,
                        "especialidad": especialidad,
                    }
                    st.session_state.rh_preview = rh_preview

                    # ✅ No borramos rh_cache ni reseteamos rh_step → memoria al volver atrás
                    st.session_state.rh_done = True
                    st.session_state.step = 6
                    st.rerun()

                except PermissionError:
                    st.error("⚠️ No se pudo guardar: el archivo está abierto con bloqueo de escritura en Excel.")
                    st.stop()
                except Exception as e:
                    st.error(f"⚠️ Error al escribir datos de Robos/Hurtos: {e}")
                    st.stop()
