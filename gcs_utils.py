import json
from typing import Any, Dict, Optional

import streamlit as st
from google.cloud import storage
from google.api_core.exceptions import NotFound


@st.cache_resource
def _get_storage_client() -> storage.Client:
    """
    Crea un cliente de GCS usando la info guardada en st.secrets["gcs_service_account"].

    Soporta dos casos:
    - Secret como string JSON:
        gcs_service_account = """{...json...}"""
    - Secret como tabla TOML:
        [gcs_service_account]
        type = "service_account"
        ...
    """
    raw_info = st.secrets["gcs_service_account"]

    # ⬇️ ESTA ES LA CLAVE: no siempre hacemos json.loads
    if isinstance(raw_info, str):
        # Caso: lo guardaste como string JSON entre comillas
        try:
            info = json.loads(raw_info)
        except Exception as exc:
            raise RuntimeError(
                "El secret 'gcs_service_account' no es un JSON válido. "
                "Revisa que hayas pegado el archivo .json completo entre las comillas."
            ) from exc
    else:
        # Caso: secret tipo tabla TOML, ya es un dict-like
        info = dict(raw_info)

    return storage.Client.from_service_account_info(info)

import io
import mimetypes

from openpyxl import Workbook, load_workbook

BUCKET_NAME = "operaciones-storage"


def _get_bucket() -> storage.Bucket:
    client = _get_storage_client()
    return client.bucket(BUCKET_NAME)


def _content_type_from_name(name: str) -> Optional[str]:
    guessed, _ = mimetypes.guess_type(name)
    return guessed


def blob_exists(blob_name: str) -> bool:
    client = _get_storage_client()
    bucket = _get_bucket()
    blob = bucket.blob(blob_name)
    return blob.exists(client=client)


def download_blob_bytes(blob_name: str) -> Optional[bytes]:
    client = _get_storage_client()
    bucket = _get_bucket()
    blob = bucket.blob(blob_name)
    if not blob.exists(client=client):
        return None
    return blob.download_as_bytes()


def upload_blob_bytes(blob_name: str, data: bytes, content_type: Optional[str] = None) -> None:
    bucket = _get_bucket()
    blob = bucket.blob(blob_name)
    blob.upload_from_string(data, content_type=content_type or _content_type_from_name(blob_name))


# =========================
# Excel helpers
# =========================

EXCEL_EXTS = (".xlsm", ".xlsx", ".xls")


def is_xlsm(name: str) -> bool:
    return name.lower().endswith(".xlsm")


def resolve_excel_blob(base_without_ext: str) -> str:
    for ext in EXCEL_EXTS:
        candidate = f"{base_without_ext}{ext}"
        if blob_exists(candidate):
            return candidate
    return f"{base_without_ext}.xlsm"


def ensure_excel_blob(blob_name: str) -> str:
    if download_blob_bytes(blob_name) is not None:
        return blob_name
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    save_workbook_to_gcs(wb, blob_name)
    return blob_name


def load_workbook_from_gcs(blob_name: str):
    excel_blob = ensure_excel_blob(blob_name)
    data = download_blob_bytes(excel_blob)
    if data is None:
        wb = Workbook()
        ws = wb.active
        ws.title = "Hoja1"
        save_workbook_to_gcs(wb, excel_blob)
        return wb
    return load_workbook(io.BytesIO(data), keep_vba=is_xlsm(excel_blob))


def save_workbook_to_gcs(wb, blob_name: str) -> None:
    buffer = io.BytesIO()
    wb.save(buffer)
    upload_blob_bytes(blob_name, buffer.getvalue(), content_type=_content_type_from_name(blob_name))


# =========================
# JSON helpers (agenda)
# =========================

def load_json_from_gcs(blob_name: str) -> dict[str, Any]:
    try:
        data = download_blob_bytes(blob_name)
    except NotFound:
        return {}

    if not data:
        return {}

    try:
        return json.loads(data.decode("utf-8"))
    except json.JSONDecodeError:
        return {}


def save_json_to_gcs(blob_name: str, payload: dict[str, Any]) -> None:
    upload_blob_bytes(blob_name, json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8"), content_type="application/json")
