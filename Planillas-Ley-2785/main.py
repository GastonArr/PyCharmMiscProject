import datetime as dt
import sys
from pathlib import Path

import streamlit as st
from openpyxl.utils import column_index_from_string

_PLANILLAS_DIR = Path(__file__).resolve().parent
_SNICSAT_DIR = _PLANILLAS_DIR.parent / "SNIC-SAT"
if str(_SNICSAT_DIR) not in sys.path:
    sys.path.insert(0, str(_SNICSAT_DIR))

from gcs_utils import (
    blob_exists,
    download_blob_bytes,
    load_workbook_from_gcs,
    resolve_excel_blob,
    save_workbook_to_gcs,
    upload_blob_bytes,
)
from agenda_ley_2785 import (
    registrar_carga_hecho,
    render_admin_agenda,
    render_selector_agenda,
    render_selector_unidad,
)

from paso1 import render_paso1
from paso2 import render_paso2
from paso3 import render_paso3
from paso4 import render_paso4

# ==========================
# CONFIGURACIÓN BÁSICA
# ==========================

_PAGE_CONFIGURED = False


def _configure_page() -> None:
    global _PAGE_CONFIGURED
    if _PAGE_CONFIGURED:
        return
    st.set_page_config(page_title="Carga LEY 2785", layout="wide")
    _PAGE_CONFIGURED = True


# Rutas en Cloud Storage
BLOB_PREFIX = "planillas-ley-2785"
TEMPLATE_BASE = f"{BLOB_PREFIX}/plantilla/PLANILLA LEY 2785 NUEVA"
EXCEL_SHEET_NAME = "LEY 2785"

# Mapear unidad -> archivo Excel de salida en el bucket
UNIT_FILE_MAP = {
    "Comisaría 6°": f"{BLOB_PREFIX}/LEY-2785-COMISARIA6.xlsx",
    "Comisaría 9°": f"{BLOB_PREFIX}/LEY-2785-COMISARIA9.xlsx",
    "Comisaría 14°": f"{BLOB_PREFIX}/LEY-2785-COMISARIA14.xlsx",
    "Comisaría 15°": f"{BLOB_PREFIX}/LEY-2785-COMISARIA15.xlsx",
    "Comisaría 42°": f"{BLOB_PREFIX}/LEY-2785-COMISARIA42.xlsx",
    "CNAF 4": f"{BLOB_PREFIX}/LEY-2785-CNAF-4.xlsx",
}
UNIDADES_JURISDICCION = list(UNIT_FILE_MAP.keys())

# ==========================
# LISTAS DE OPCIONES
# ==========================

DOCUMENTO_OPTIONS = [
    "DNI/DU",
    "LC",
    "LE",
    "CI",
    "Pasaporte Extranjero",
    "Otro documento",
    "No informado",
]

SEXO1_OPTIONS = ["Mujer", "Varon", "No Informado"]
TRANS1_OPTIONS = ["Travesti", "Transexual", "Otra", "No informado"]

EDUCACION1_OPTIONS = [
    "Sin instrucción",
    "Primario",
    "EGB",
    "Polimodal",
    "Secundario",
    "Terciario",
    "Universitario",
    "Educación especial",
    "No sabe",
    "No informado",
]

COMPLITUD1_OPTIONS = ["SI", "NO", "NO SABE", "NO INFORMADO"]

OCUPADA1_OPTIONS = ["Ocupada/o", "No ocupada/o", "NO INFORMADO"]

ACTIVIDAD1_OPTIONS = [
    "estudiante",
    "jubilada/o-pensión",
    "Ama de casa",
    "otra",
    "No informado",
]

VINCULO_OPTIONS = [
    "Pareja/novio",
    "ex pareja",
    "padre",
    "madre",
    "hijo",
    "hija",
    "otros",
    "desconocido",
]

CONVIVENCIA_OPTIONS = ["si", "no", "no sabe", "no informado"]

TIPO_OPTIONS = ["SI", "NO"]

MODALIDAD_OPTIONS = [
    "Doméstica",
    "Institucional",
    "Laboral",
    "Contra la libertad reproductiva",
    "Obstétrica",
    "No informado",
]

TIEMPO_OPTIONS = [
    "Menos de un año",
    "de 1 a 5 años",
    "de 6 a 10 años",
    "más de 10 años",
    "No recureda",
    "No informado",
]

FRECUENCIA_OPTIONS = ["Sólo una vez", "Más de una vez", "No informado"]

SEXO2_OPTIONS = ["Mujer", "Varón", "No Sabe", "No informado"]
TRANS2_OPTIONS = ["Travesti", "Transexual", "Otra", "No informado"]

EDUCACION2_OPTIONS = list(EDUCACION1_OPTIONS)
COMPLITUD2_OPTIONS = ["Si", "No", "No Sabe", "No informado"]
ACTIVIDAD2_OPTIONS = ["Ocupada/o", "No ocupada/o", "No informado"]

OTRA2_OPTIONS = [
    "estudiante",
    "jubilada/o-pensión",
    "Ama de casa",
    "otra",
    "No informado",
]

# ==========================
# COLUMNAS
# ==========================

COLUMN_MAPPING = {
    "tipo_documento": "C",
    "otro_doc": "D",  # sigue existiendo, pero no se carga en el formulario
    "identificacion": "E",
    "institucion": "F",
    "fecha_consulta": "G",
    "sexo1": "H",
    "trans1": "I",
    "edad": "J",
    "provincia": "K",
    "partido_municipio": "L",
    "localidad": "M",
    "nivel_educativo1": "N",
    "complitud1": "O",
    "ocupada1": "P",
    "actividad1": "Q",
    "vinculo": "R",
    "otro_vinculo": "S",
    "convivencia": "T",
    "viol_fisica": "U",
    "viol_psico": "V",
    "viol_econ": "W",
    "viol_sexual": "X",
    "modalidad": "Y",
    "tiempo": "Z",
    "frecuencia": "AA",
    "sexo2": "AB",
    "trans2": "AC",
    "edad_agresor": "AD",
    "nivel_educativo2": "AE",
    "complitud2": "AF",
    "actividad2": "AG",
    "otra_actividad2": "AH",
    "info_especifica": "AI",
    "fecha_modificacion": "AJ",
}

FIELD_LABELS = {
    "tipo_documento": "Tipo de documento (columna C)",
    "identificacion": "Identificación (columna E)",
    "institucion": "Unidad / Institución (columna F)",
    "fecha_consulta": "Fecha de consulta (columna G)",
    "sexo1": "Sexo (columna H)",
    "trans1": "Identidad trans (columna I)",
    "edad": "Edad (columna J)",
    "provincia": "Provincia (columna K)",
    "localidad": "Localidad (columna M)",
    "nivel_educativo1": "Nivel educativo (columna N)",
    "complitud1": "Complitud educativo (columna O)",
    "ocupada1": "Situación ocupacional (columna P)",
    "actividad1": "Actividad (columna Q)",
    "vinculo": "Vínculo (columna R)",
    "otro_vinculo": "Otro vínculo (columna S)",
    "convivencia": "Convivencia (columna T)",
    "viol_fisica": "Violencia física (columna U)",
    "viol_psico": "Violencia psicológica (columna V)",
    "viol_econ": "Violencia económica (columna W)",
    "viol_sexual": "Violencia sexual (columna X)",
    "modalidad": "Modalidad (columna Y)",
    "tiempo": "Tiempo del maltrato (columna Z)",
    "frecuencia": "Frecuencia (columna AA)",
    "sexo2": "Sexo agresor (columna AB)",
    "trans2": "Trans agresor (columna AC)",
    "edad_agresor": "Edad agresor (columna AD)",
    "nivel_educativo2": "Nivel educativo agresor (columna AE)",
    "complitud2": "Complitud agresor (columna AF)",
    "actividad2": "Actividad laboral agresor (columna AG)",
    "otra_actividad2": "Otra actividad agresor (columna AH)",
    "info_especifica": "Información específica (columna AI)",
}

STEP_REQUIRED = {
    1: ["institucion", "fecha_consulta", "tipo_documento", "identificacion"],
    2: [
        "sexo1",
        "trans1",
        "edad",
        "provincia",
        "localidad",
        "nivel_educativo1",
        "complitud1",
        "ocupada1",
        "actividad1",
    ],
    3: [
        "vinculo",
        "otro_vinculo",
        "convivencia",
        "viol_fisica",
        "viol_psico",
        "viol_econ",
        "viol_sexual",
        "modalidad",
        "tiempo",
        "frecuencia",
    ],
    4: [
        "sexo2",
        "trans2",
        "edad_agresor",
        "nivel_educativo2",
        "complitud2",
        "actividad2",
        "otra_actividad2",
        "info_especifica",
    ],
}

REQUIRED_FIELDS = sorted({f for step in STEP_REQUIRED.values() for f in step})
ZERO_NOT_ALLOWED_FIELDS = {"edad", "edad_agresor"}

# ==========================
# FUNCIONES AUXILIARES
# ==========================

def initialize_default_state():
    """
    No seteamos institucion, tipo_documento ni identificacion acá,
    porque se manejan en paso1.py según lo que elija el usuario.
    """
    defaults = {
        "fecha_consulta": dt.date.today(),
        "sexo1": SEXO1_OPTIONS[0],
        "trans1": TRANS1_OPTIONS[0],
        "edad": 0,
        "nivel_educativo1": EDUCACION1_OPTIONS[0],
        "complitud1": COMPLITUD1_OPTIONS[0],
        "ocupada1": OCUPADA1_OPTIONS[0],
        "actividad1": ACTIVIDAD1_OPTIONS[0],
        "vinculo": VINCULO_OPTIONS[0],
        "convivencia": CONVIVENCIA_OPTIONS[0],
        "viol_fisica": TIPO_OPTIONS[0],
        "viol_psico": TIPO_OPTIONS[0],
        "viol_econ": TIPO_OPTIONS[0],
        "viol_sexual": TIPO_OPTIONS[0],
        "modalidad": MODALIDAD_OPTIONS[0],
        "tiempo": TIEMPO_OPTIONS[0],
        "frecuencia": FRECUENCIA_OPTIONS[0],
        "sexo2": SEXO2_OPTIONS[0],
        "trans2": TRANS2_OPTIONS[0],
        "edad_agresor": 0,
        "nivel_educativo2": EDUCACION2_OPTIONS[0],
        "complitud2": COMPLITUD2_OPTIONS[0],
        "actividad2": ACTIVIDAD2_OPTIONS[0],
        "otra_actividad2": OTRA2_OPTIONS[0],
        "info_especifica": "",
        "otro_vinculo": "",
        "fecha_modificacion": "",
    }
    for k, v in defaults.items():
        st.session_state.setdefault(k, v)


def sanitize_required_text_fields():
    """
    Limpia texto obligatorio de espacios y None.
    Incluye identificacion, provincia, localidad, otro_doc, otro_vinculo.
    """
    out = {}
    for k in ["identificacion", "provincia", "localidad", "otro_doc", "otro_vinculo"]:
        v = st.session_state.get(k, "")
        if isinstance(v, str):
            v = v.strip()
        elif v is None:
            v = ""
        out[k] = v
    return out


def build_form_data_from_state():
    clean = sanitize_required_text_fields()
    out = {}
    for k in COLUMN_MAPPING.keys():
        if k == "partido_municipio":
            val = clean.get("localidad", st.session_state.get("localidad"))
        else:
            val = clean.get(k, st.session_state.get(k))

        if k == "fecha_consulta" and isinstance(val, dt.date):
            val = val.strftime("%d/%m/%Y")

        out[k] = val
    return out


def find_missing_in_state(keys):
    clean = sanitize_required_text_fields()
    missing = []

    for k in keys:
        val = clean.get(k, st.session_state.get(k))

        if val is None:
            missing.append(k)
            continue

        if isinstance(val, str) and val.strip() == "":
            missing.append(k)
            continue

        if k in ZERO_NOT_ALLOWED_FIELDS:
            try:
                if int(val) == 0:
                    missing.append(k)
            except Exception:
                missing.append(k)

    return missing


def ensure_unit_file_exists(unidad):
    if unidad not in UNIT_FILE_MAP:
        raise ValueError(f"Unidad no reconocida: {unidad}")

    target = UNIT_FILE_MAP[unidad]

    template_blob = resolve_excel_blob(TEMPLATE_BASE)
    template_bytes = download_blob_bytes(template_blob)
    if template_bytes is None:
        raise FileNotFoundError(
            "No se encontró la plantilla base en el almacenamiento. "
            f"Verificá que exista el blob '{template_blob}'."
        )

    if not blob_exists(target):
        upload_blob_bytes(target, template_bytes)

    return target


def get_next_row_and_counter(ws):
    row = 3
    while ws.cell(row=row, column=1).value not in (None, ""):
        row += 1

    if row == 3:
        return row, 1

    last = ws.cell(row=row - 1, column=1).value
    try:
        last = int(last)
    except Exception:
        last = 0

    return row, last + 1


def save_to_excel(unidad, data):
    target = ensure_unit_file_exists(unidad)
    wb = load_workbook_from_gcs(target)
    if EXCEL_SHEET_NAME not in wb.sheetnames:
        raise ValueError(
            f"La hoja '{EXCEL_SHEET_NAME}' no existe en el archivo {target}."
        )

    ws = wb[EXCEL_SHEET_NAME]

    row, counter = get_next_row_and_counter(ws)
    ws.cell(row=row, column=1).value = counter

    for k, col in COLUMN_MAPPING.items():
        idx = column_index_from_string(col)
        ws.cell(row=row, column=idx).value = data.get(k)
    save_workbook_to_gcs(wb, target)
    return counter, target


def render_admin_download(unidades):
    st.sidebar.markdown("---")
    st.sidebar.subheader("Descarga de planillas")

    try:
        default_idx = unidades.index(st.session_state.get("institucion", unidades[0]))
    except ValueError:
        default_idx = 0

    unidad_descarga = st.sidebar.selectbox(
        "Seleccione la comisaría",
        unidades,
        index=default_idx,
        key="unidad_descarga_admin",
    )

    try:
        blob_name = ensure_unit_file_exists(unidad_descarga)
    except FileNotFoundError as exc:
        st.sidebar.error(str(exc))
        return

    data = download_blob_bytes(blob_name)
    if not data:
        st.sidebar.warning(
            "No se pudo obtener la planilla seleccionada desde el almacenamiento en la nube."
        )
        return

    st.sidebar.download_button(
        "Descargar planilla de la comisaría seleccionada",
        data=data,
        file_name=Path(blob_name).name,
        use_container_width=True,
    )


def reset_form():
    for k in COLUMN_MAPPING.keys():
        st.session_state.pop(k, None)
    st.session_state.step = 1


def _referencia_banner(etiqueta_hecho: str, referencia: str | None) -> None:
    ref = referencia.strip() if isinstance(referencia, str) else ""
    titulo = etiqueta_hecho or "Hecho asignado"
    if ref:
        st.info(
            f"{titulo}\n\nInformación específica (AI): {ref}",
            icon="🗂️",
        )
    else:
        st.warning(
            f"{titulo}\n\nSin referencia de 'Información específica (AI)'.",
            icon="ℹ️",
        )


def _allowed_units(unidades):
    if not unidades:
        return UNIDADES_JURISDICCION
    filtered = [u for u in unidades if u in UNIDADES_JURISDICCION]
    return filtered


def run_planillas_ley_2785_app(allowed_units=None, configure_page=True, is_admin=False):
    if configure_page:
        _configure_page()

    unidades = _allowed_units(allowed_units)
    if not unidades:
        st.error(
            "No cuenta con unidades habilitadas para cargar planillas Ley 2785."
        )
        return

    if is_admin:
        render_admin_download(unidades)
        render_admin_agenda(st.session_state.get("username"), unidades)

    if "step" not in st.session_state:
        st.session_state.step = 1

    initialize_default_state()

    st.title("Carga de registros - Ley 2785")

    unidad_sel = render_selector_unidad(unidades)
    fecha_agenda, hechos_pendientes, mensaje_agenda = render_selector_agenda(unidad_sel)

    if mensaje_agenda:
        if str(mensaje_agenda).startswith("¡Felicitaciones!"):
            st.success(mensaje_agenda)
        else:
            st.warning(mensaje_agenda)
        if not fecha_agenda:
            return

    if not fecha_agenda or not hechos_pendientes:
        st.info("No hay hechos planificados para continuar con la carga.")
        return

    hechos_ids = list(hechos_pendientes.keys())
    default_hecho = st.session_state.get("agenda_planillas_hecho")
    if default_hecho not in hechos_ids:
        default_hecho = hechos_ids[0]

    def _format_hecho(hecho_id):
        info = hechos_pendientes.get(hecho_id, {})
        return info.get("display") or info.get("etiqueta") or hecho_id

    hecho_sel = st.selectbox(
        "Seleccione el hecho asignado",
        options=hechos_ids,
        format_func=_format_hecho,
        index=hechos_ids.index(default_hecho),
        key="agenda_planillas_hecho",
    )

    info_hecho = hechos_pendientes.get(hecho_sel, {})
    referencia_actual = info_hecho.get("referencia") or ""
    etiqueta_hecho = info_hecho.get("display") or info_hecho.get("etiqueta") or "Hecho asignado"

    if st.session_state.get("agenda_planillas_hecho_prev") != hecho_sel:
        st.session_state.step = 1
    st.session_state["agenda_planillas_hecho_prev"] = hecho_sel

    st.session_state["info_especifica"] = referencia_actual
    st.session_state["agenda_planillas_etiqueta"] = etiqueta_hecho
    st.session_state["agenda_planillas_fecha"] = fecha_agenda
    st.session_state["institucion"] = unidad_sel

    _referencia_banner(etiqueta_hecho, referencia_actual)

    steps_total = 4
    st.progress(st.session_state.step / steps_total)
    st.caption(f"Paso {st.session_state.step} de {steps_total}")

    agenda_context = {
        "fecha": fecha_agenda,
        "hecho_id": hecho_sel,
        "unidad": unidad_sel,
        "referencia": referencia_actual,
        "registrar": registrar_carga_hecho,
    }

    if st.session_state.step == 1:
        render_paso1(unidades, DOCUMENTO_OPTIONS)

    elif st.session_state.step == 2:
        render_paso2(
            SEXO1_OPTIONS,
            TRANS1_OPTIONS,
            EDUCACION1_OPTIONS,
            COMPLITUD1_OPTIONS,
            OCUPADA1_OPTIONS,
            ACTIVIDAD1_OPTIONS,
        )

    elif st.session_state.step == 3:
        render_paso3(
            VINCULO_OPTIONS,
            CONVIVENCIA_OPTIONS,
            TIPO_OPTIONS,
            MODALIDAD_OPTIONS,
            TIEMPO_OPTIONS,
            FRECUENCIA_OPTIONS,
        )

    elif st.session_state.step == 4:
        render_paso4(
            SEXO2_OPTIONS,
            TRANS2_OPTIONS,
            EDUCACION2_OPTIONS,
            COMPLITUD2_OPTIONS,
            ACTIVIDAD2_OPTIONS,
            OTRA2_OPTIONS,
            STEP_REQUIRED,
            REQUIRED_FIELDS,
            FIELD_LABELS,
            find_missing_in_state,
            build_form_data_from_state,
            save_to_excel,
            reset_form,
            agenda_context,
        )

    st.markdown("---")
    col1, _, col3 = st.columns(3)

    with col1:
        if st.session_state.step > 1:
            if st.button("⬅ Anterior"):
                st.session_state.step -= 1
                st.rerun()

    with col3:
        if st.session_state.step < steps_total:
            if st.button("Siguiente ➡"):
                required = STEP_REQUIRED[st.session_state.step]
                missing = find_missing_in_state(required)

                if missing:
                    labels = [FIELD_LABELS[k] for k in missing]
                    st.error(
                        "Para continuar, completá los siguientes campos:\n\n- "
                        + "\n- ".join(labels)
                    )
                else:
                    st.session_state.step += 1
                    st.rerun()


if __name__ == "__main__":
    run_planillas_ley_2785_app()
