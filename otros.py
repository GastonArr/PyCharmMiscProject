# otros.py
# Subflujo “Otros” para:
# - LESIONES GRAVES / LESIONES LEVES / LESIONES GRAVISIMAS / OTROS DELITOS CONTRA LA LIBERTAD INDIVIDUAL
# - ABUSO DE ARMAS CON LESIONES
# - ABUSO DE ARMAS SIN LESIONES
# - ABUSO SEXUAL CON ACCESO CARNAL (VIOLACION)
# - ABUSO SEXUAL SIMPLE
#     · Víctimas por sexo → AO (total) y AG/AH/AI
#     · VULNERABILIDAD DE LA VÍCTIMA (AP)
# - DESAPARICION DE PERSONA:
#     · Lo mismo que arriba + “¿Apareció?” (BA: SI/NO)
#
# Integración:
#   import otros
#   otros.render(excel_path=st.session_state.excel_path, fila=st.session_state.fila, delito_x3=st.session_state.delito)

import streamlit as st
from openpyxl import load_workbook, Workbook
import os
import cloud_storage

# ==============================
# Utilidades de Excel
# ==============================
def is_xlsm(path: str) -> bool:
    return path.lower().endswith(".xlsm")

def asegurar_excel(path: str):
    cloud_storage.ensure_local_file(path)
    carpeta = os.path.dirname(path)
    if carpeta and not os.path.exists(carpeta):
        os.makedirs(carpeta, exist_ok=True)
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Hoja1"
        wb.save(path)
        cloud_storage.sync_local_to_remote(path)

def cargar_libro(path: str):
    asegurar_excel(path)
    return load_workbook(path, keep_vba=is_xlsm(path))

def unwrap_quotes(v):
    """Quita SOLO comillas envolventes si existen; no modifica espacios internos."""
    if not isinstance(v, str):
        return v
    if len(v) >= 2 and ((v[0] == v[-1] == '"') or (v[0] == v[-1] == "'")):
        return v[1:-1]
    return v

def _trim(s):
    return s.strip() if isinstance(s, str) else s

def mostrar_hecho_referencia():
    """Muestra el hecho cargado en el flujo principal para tenerlo siempre a la vista."""
    hecho = st.session_state.get("hecho")
    if hecho:
        st.subheader("Hecho ingresado para referencia:")
        st.write(hecho)
        st.markdown("---")

# ==============================
# Delitos que activan este flujo
# ==============================
# Importante: usamos .strip() al comparar, para tolerar espacios finales en la lista principal.
DELITOS_LESIONES = {
    "LESIONES GRAVES",
    "LESIONES LEVES",
    "LESIONES GRAVISIMAS",
    "OTROS DELITOS CONTRA LA LIBERTAD INDIVIDUAL",
}

DELITOS_ABUSO_ARMAS = {
    "ABUSO DE ARMAS CON LESIONES",
    "ABUSO DE ARMAS SIN LESIONES",
}

DELITOS_SEXUALES = {
    "ABUSO SEXUAL CON ACCESO CARNAL (VIOLACION)",
    "ABUSO SEXUAL SIMPLE",
}

DELITO_DESAPARICION = "DESAPARICION DE PERSONA"

# ==============================
# Catálogos
# ==============================
VULNERABILIDAD = [
    "No pertenece a ningun grupo vulnerable",
    "Niñez - Edad joven",
    "Mujer embarazada",
    "Persona con discapacidad",
    "Personas en situacion de calle",
    "Migrantes",
    "Minorías Etnicas, Raciales o Religiosas",
    "Adultos Mayores",
    "Sin determinar",
]

SEXO_OPCIONES = ["MASCULINO", "FEMENINO", "NO CONSTA"]
CANT_1_10 = [str(i) for i in range(1, 11)]
SI_NO = ["SI", "NO"]

# ==============================
# Render principal
# ==============================
def render(excel_path: str, fila: int, delito_x3: str) -> None:
    """
    Manejo de estado: st.session_state.others_step, others_done, others_preview.
    Con este cambio NO se escribe en Excel aquí: solo se valida, se arma others_preview
    y se vuelve al resumen (step=6). El guardado real va en “Finalizar y guardar ✅”.
    """
    delito_norm = (delito_x3 or "").strip()

    aplica = (
        delito_norm in DELITOS_LESIONES
        or delito_norm in DELITOS_ABUSO_ARMAS
        or delito_norm in DELITOS_SEXUALES
        or delito_norm == DELITO_DESAPARICION
    )
    if not aplica:
        st.session_state.step = 6
        st.rerun()

    if "others_step" not in st.session_state:
        st.session_state.others_step = 1

    # === NUEVO: precarga desde others_preview para conservar datos al volver ===
    ss = st.session_state
    prev = ss.get("others_preview") or {}

    # Sembrar lista dinámica de víctimas si viene desde preview y aún no existe
    if ("others_vict_rows" not in ss or not isinstance(ss.others_vict_rows, list)) and prev.get("vict_rows"):
        ss.others_vict_rows = list(prev.get("vict_rows"))

    # Helper para no pisar lo recién tipeado
    def seed(k, v):
        if v not in (None, "") and k not in ss:
            ss[k] = v

    # Sembrar valores simples para widgets
    seed("others_vulnerab", prev.get("vulnerabilidad"))
    if delito_norm == DELITO_DESAPARICION:
        seed("others_aparecio", prev.get("aparecio"))

    def C(col: str) -> str:
        return f"{col}{fila}"

    # =============== Paso único: Captura (sin guardar acá) =================
    if st.session_state.others_step == 1:
        st.subheader("Datos adicionales")
        mostrar_hecho_referencia()

        # ---------- Víctimas: distribución por sexo (dinámica) ----------
        st.markdown("### Víctimas")
        st.caption("Distribución por sexo de las víctimas. Podés agregar otro sexo si corresponde (máx. 3 filas, total hasta 10).")

        if "others_vict_rows" not in st.session_state or not isinstance(st.session_state.others_vict_rows, list):
            st.session_state.others_vict_rows = [{"sexo": "MASCULINO", "cant": "1"}]

        total_vict = 0
        vict_borrar = []
        for i, row in enumerate(st.session_state.others_vict_rows):
            c1, c2, c3 = st.columns([2, 1, 1])
            with c1:
                vict_sexo = st.selectbox(
                    "Sexo",
                    SEXO_OPCIONES,
                    index=SEXO_OPCIONES.index(row.get("sexo", "MASCULINO")) if row.get("sexo") in SEXO_OPCIONES else 0,
                    key=f"others_vict_rows_{i}_sexo",
                )
            with c2:
                vict_cant = st.selectbox(
                    "Cantidad",
                    CANT_1_10,
                    index=(CANT_1_10.index(row.get("cant")) if row.get("cant") in CANT_1_10 else 0),
                    key=f"others_vict_rows_{i}_cant",
                )
            with c3:
                if i > 0:
                    if st.button("Quitar", key=f"others_vict_rows_{i}_del"):
                        vict_borrar.append(i)

            st.session_state.others_vict_rows[i]["sexo"] = vict_sexo
            st.session_state.others_vict_rows[i]["cant"] = vict_cant
            total_vict += int(vict_cant)

        if vict_borrar:
            st.session_state.others_vict_rows = [r for j, r in enumerate(st.session_state.others_vict_rows) if j not in vict_borrar]
            st.rerun()

        if len(st.session_state.others_vict_rows) < 3 and total_vict < 10:
            if st.button("Agregar otro sexo", key="others_add_vict_row"):
                usados = {r["sexo"] for r in st.session_state.others_vict_rows if "sexo" in r}
                restantes = [s for s in SEXO_OPCIONES if s not in usados]
                nuevo_sexo = restantes[0] if restantes else SEXO_OPCIONES[0]
                st.session_state.others_vict_rows.append({"sexo": nuevo_sexo, "cant": "1"})
                st.rerun()

        st.caption(f"Total víctimas (AO): {total_vict}")

        st.markdown("---")

        # ---------- Vulnerabilidad ----------
        vulnerab_sel = st.selectbox("VULNERABILIDAD DE LA VICTIMA (AP)", VULNERABILIDAD, key="others_vulnerab")

        # ---------- ¿Apareció? (solo Desaparición) ----------
        aparecio_sel = None
        if delito_norm == DELITO_DESAPARICION:
            aparecio_sel = st.radio("Consignar si APARECIO (BA)", SI_NO, horizontal=True, key="others_aparecio")

        st.markdown("---")

        # Navegación
        col1, col2 = st.columns(2)
        with col1:
            if st.button("Volver ⬅️", key="others_volver"):
                # ⬅️ ARREGLADO: volver a Direcciones (step 5) en lugar de step 6
                st.session_state.step = 5
                st.rerun()
        with col2:
            if st.button("Guardar y continuar ➡️", key="others_guardar"):
                # Validaciones (idénticas a las tuyas)
                if total_vict < 1:
                    st.warning("Debe indicar al menos 1 víctima.")
                    st.stop()
                if total_vict > 10:
                    st.warning("El total de víctimas no puede superar 10.")
                    st.stop()
                if not vulnerab_sel:
                    st.warning("Debe seleccionar la Vulnerabilidad (AP).")
                    st.stop()
                if delito_norm == DELITO_DESAPARICION and not aparecio_sel:
                    st.warning("Debe consignar si APARECIÓ (BA).")
                    st.stop()

                # (DIFERIDO) No escribir en Excel acá — solo armar preview y avanzar
                try:
                    # Acumular por sexo para el total (AO)
                    total_m = total_f = total_nc = 0
                    for r in st.session_state.others_vict_rows:
                        sexo = (r.get("sexo") or "").strip()
                        try:
                            c = int(r.get("cant") or "0")
                        except Exception:
                            c = 0
                        if sexo == "MASCULINO":
                            total_m += c
                        elif sexo == "FEMENINO":
                            total_f += c
                        elif sexo == "NO CONSTA":
                            total_nc += c
                    ao_total = total_m + total_f + total_nc

                    # Preview para el resumen del paso 6
                    st.session_state.others_preview = {
                        "vict_rows": list(st.session_state.others_vict_rows),
                        "vict_total": str(ao_total),
                        "vulnerabilidad": vulnerab_sel,
                        "aparecio": (aparecio_sel if delito_norm == DELITO_DESAPARICION else None),
                    }

                    # Fin subflujo (sin borrar listas ni step, así “Atrás” conserva todo)
                    st.session_state.others_done = True
                    st.session_state.step = 6
                    st.rerun()

                except PermissionError:
                    st.error("⚠️ No se pudo guardar: el archivo está abierto con bloqueo de escritura en Excel.")
                    st.stop()
                except Exception as e:
                    st.error(f"⚠️ Error al escribir datos: {e}")
                    st.stop()