import json
import os
import tempfile
from io import BytesIO
from typing import Optional

import pandas as pd
import streamlit as st
from google.cloud import storage
from openpyxl import Workbook, load_workbook

GCS_BUCKET_NAME = "proyecto-operaciones-storage"


def get_gcs_client() -> storage.Client:
    """
    Crea un cliente de Google Cloud Storage usando las credenciales guardadas
    en st.secrets["gcs_service_account"], que es un JSON en formato string.
    """
    raw_info = st.secrets["gcs_service_account"]
    info = json.loads(raw_info)
    return storage.Client.from_service_account_info(info)


def get_bucket():
    client = get_gcs_client()
    return client.bucket(GCS_BUCKET_NAME)


def download_blob_to_tempfile(blob_name: str, suffix: str = "") -> str:
    """
    Descarga un blob del bucket a un archivo temporal local y devuelve la ruta.
    Se usa cuando alguna parte del código necesita un path de archivo .xlsx.
    """
    bucket = get_bucket()
    blob = bucket.blob(blob_name)

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    blob.download_to_filename(tmp.name)
    return tmp.name


def read_excel_from_gcs(blob_name: str, **read_excel_kwargs) -> pd.DataFrame:
    """
    Lee un Excel directamente desde GCS a un DataFrame de pandas sin necesidad
    de guardar un archivo permanente en disco.
    """
    bucket = get_bucket()
    blob = bucket.blob(blob_name)
    data = blob.download_as_bytes()
    bio = BytesIO(data)
    return pd.read_excel(bio, **read_excel_kwargs)


def write_excel_to_gcs(df: pd.DataFrame, blob_name: str, **to_excel_kwargs) -> None:
    """
    Escribe un DataFrame de pandas como Excel en un blob de GCS, sobrescribiendo
    el contenido anterior.
    """
    bucket = get_bucket()
    blob = bucket.blob(blob_name)

    bio = BytesIO()
    df.to_excel(bio, index=False, **to_excel_kwargs)
    bio.seek(0)

    blob.upload_from_file(
        bio,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# Utilidades específicas para archivos Excel gestionados con openpyxl


def is_xlsm(name: str) -> bool:
    return name.lower().endswith(".xlsm")


def _default_blob_name(base_without_ext: str) -> str:
    return f"{base_without_ext}.xlsm"


def resolve_excel_blob_name(base_without_ext: str) -> str:
    bucket = get_bucket()
    for ext in (".xlsm", ".xlsx", ".xls"):
        cand = f"{base_without_ext}{ext}"
        if bucket.blob(cand).exists():
            return cand
    return _default_blob_name(base_without_ext)


def ensure_excel_blob(blob_name: str) -> None:
    bucket = get_bucket()
    blob = bucket.blob(blob_name)
    if blob.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "Hoja1"
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    content_type = (
        "application/vnd.ms-excel.sheet.macroEnabled.12"
        if is_xlsm(blob_name)
        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    blob.upload_from_file(bio, content_type=content_type)


def load_workbook_from_gcs(blob_name: str):
    ensure_excel_blob(blob_name)
    bucket = get_bucket()
    blob = bucket.blob(blob_name)
    data = blob.download_as_bytes()
    bio = BytesIO(data)
    return load_workbook(bio, keep_vba=is_xlsm(blob_name))


def save_workbook_to_gcs(wb, blob_name: str) -> None:
    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)
    content_type = (
        "application/vnd.ms-excel.sheet.macroEnabled.12"
        if is_xlsm(blob_name)
        else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    bucket = get_bucket()
    blob = bucket.blob(blob_name)
    blob.upload_from_file(bio, content_type=content_type)


def blob_exists(blob_name: str) -> bool:
    return get_bucket().blob(blob_name).exists()
