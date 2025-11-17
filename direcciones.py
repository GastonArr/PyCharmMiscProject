import streamlit as st
from openpyxl import load_workbook, Workbook
import os

# ===== Utilidades Excel =====
def is_xlsm(path: str) -> bool:
    return path.lower().endswith(".xlsm")

def asegurar_excel(path: str):
    carpeta = os.path.dirname(path)
    if carpeta and not os.path.exists(carpeta):
        os.makedirs(carpeta, exist_ok=True)
    if not os.path.exists(path):
        wb = Workbook()
        ws = wb.active
        ws.title = "Hoja1"
        wb.save(path)

def cargar_libro(path: str):
    asegurar_excel(path)
    return load_workbook(path, keep_vba=is_xlsm(path))

def unwrap_quotes(v):
    """Quita SOLO comillas envolventes si existen; no toca espacios."""
    if not isinstance(v, str):
        return v
    if len(v) >= 2 and ((v[0] == v[-1] == '"') or (v[0] == v[-1] == "'")):
        return v[1:-1]
    return v

# ===== Lógica de direcciones (UI, navegación, guardado diferido) =====
def render_direcciones_ui(excel_path: str, fila: int, comisaria: str) -> None:
    """
    Pantalla de direcciones integrada al flujo principal (usa st.session_state.step).

    NO escribe en Excel acá. Solo valida y guarda un preview en:
        st.session_state.direcciones_preview = {
            "ciudad_cod", "barrio", "otro_barrio", "direccion", "altura", "link_maps"
        }
    Esos campos se vuelcan en app.py al presionar “Finalizar y guardar ✅”:
        I=ciudad_cod, J=barrio, K=otro_barrio (si 'OTRO'), L=direccion, M=altura, N=link_maps
    """

    # 1) Lista de barrios según comisaría (sin cambiar tus listas)
    barrios_1415 = [
        "PARQUE INDUSTRIAL","PARQUE OESTE","PARQUE ESTE","FILLI DEI","ZANI","SAN MARTIN",
        "AEROPARQUE","COOPERATIVA","D. SAEZ","PAMPA","CENTRO","RUCA QUIMEY","NEUQUEN CHE",
        "BELGRANO","PENI TRAPUN","BRENTANA","PROGRESO","UNION","25 DE MAYO",
        "PUEBLO NUEVO","MONTE HERMOSO C","OTRO"
    ]
    barrios_6 = [
        "25 DE MAYO","ALTOS DEL SUR","BOSCONI","CENTENARIO","CENTRAL","NOROESTE","NORTE",
        "OTANO","PARUQE INDUSTRIAL","SOUFAL","UNIVERSITARIO","UNO","ZONA INDUSTRIAL","OTRO"
    ]

    def combinar_listas_unicas(*listas):
        vistos = set()
        combinada = []
        for lista in listas:
            for item in lista:
                if item not in vistos:
                    combinada.append(item)
                    vistos.add(item)
        return combinada

    barrios_cenaf_4 = combinar_listas_unicas(barrios_1415, barrios_6)
    barrios_solo_otro = ["OTRO"]

    if comisaria in ("Comisaria 14", "Comisaria 15"):
        lista_barrios = barrios_1415
        ciudad_cod = "CUTRAL_CO"
    elif comisaria == "Comisaria 6":
        lista_barrios = barrios_6
        ciudad_cod = "PLAZA_HUINCUL"
    elif comisaria == "CENAF 4":
        lista_barrios = barrios_cenaf_4
        ciudad_cod = "CUTRAL_CO"
    elif comisaria in ("Comisaria 42", "Comisaria 9"):
        lista_barrios = barrios_solo_otro
        ciudad_cod = "CUTRAL_CO"
    else:
        lista_barrios = barrios_6
        ciudad_cod = "PLAZA_HUINCUL"

    # === Precargar widgets desde el preview si existen (no pisa lo recién tipeado) ===
    ss = st.session_state
    dprev = ss.get("direcciones_preview") or {}

    def seed(k, v):
        if v not in (None, "") and k not in ss:
            ss[k] = v

    seed("dir_barrio_select", dprev.get("barrio"))
    seed("dir_otro_barrio", dprev.get("otro_barrio"))
    seed("dir_direccion", dprev.get("direccion"))
    seed("dir_altura", dprev.get("altura"))
    seed("dir_link_maps", dprev.get("link_maps"))

    # Barrio (usar index en base al valor ya guardado)
    barrio_val = ss.get("dir_barrio_select")
    idx_barrio = lista_barrios.index(barrio_val) if barrio_val in lista_barrios else 0
    barrio = st.selectbox("Ingrese el barrio", lista_barrios, index=idx_barrio, key="dir_barrio_select")

    # “OTRO” barrio (con valor precargado si corresponde)
    otro_barrio = ""
    if barrio == "OTRO":
        otro_barrio = st.text_input(
            "Especifique otro barrio (máx 15)",
            value=ss.get("dir_otro_barrio", ""),
            max_chars=40,
            key="dir_otro_barrio",
        )

    st.markdown("---")

    # 2) Dirección y altura (con valores precargados)
    col1, col2 = st.columns(2)
    with col1:
        direccion = st.text_input(
            "Ingrese la dirección",
            value=ss.get("dir_direccion", ""),
            max_chars=70,
            key="dir_direccion",
        )
    with col2:
        altura = st.text_input(
            "Ingrese la altura",
            value=ss.get("dir_altura", ""),
            max_chars=70,
            key="dir_altura",
        )

    # 2.b) Link de Google Maps (OBLIGATORIO, con valor precargado)
    st.markdown("---")
    link_maps = st.text_input(
        "ingresa a continuación el link de la página google maps donde se ubicó el punto geográfico de la dirección dada anteriormente.",
        value=ss.get("dir_link_maps", ""),
        max_chars=500,
        key="dir_link_maps",
        placeholder="https://maps.google.com/..."
    )

    st.markdown("---")

    # 3) Botones (controlan el flujo de la app principal)
    colA, colB = st.columns(2)
    with colA:
        if st.button("Volver ⬅️", key="dir_volver_btn"):
            st.session_state.step = 4  # vuelve a "día y hora" en app.py
            st.rerun()

    with colB:
        if st.button("Siguiente ➡️", key="dir_siguiente_btn"):
            # Validaciones mínimas (idénticas a las que tenías)
            if not barrio:
                st.warning("Seleccione un barrio.")
                st.stop()
            if barrio == "OTRO" and not otro_barrio:
                st.warning("Ingrese el nombre del barrio en 'OTRO'.")
                st.stop()
            if not direccion or direccion.strip() == "":
                st.warning("Ingrese la dirección.")
                st.stop()
            if not altura or altura.strip() == "":
                st.warning("Ingrese la altura.")
                st.stop()
            # Campo obligatorio: link de Google Maps
            if not link_maps or link_maps.strip() == "":
                st.warning("Debe ingresar el link de Google Maps.")
                st.stop()

            # Guardado diferido: solo preview en session_state (NO Excel acá)
            try:
                st.session_state.direcciones_preview = {
                    "ciudad_cod": ciudad_cod,
                    "barrio": barrio,
                    "otro_barrio": (otro_barrio or "") if barrio == "OTRO" else "",
                    "direccion": direccion,
                    "altura": altura,
                    "link_maps": link_maps.strip(),
                }

                # Continuar con la app (Paso 6 en app.py: resumen + guardar)
                st.session_state.step = 6
                st.rerun()

            except PermissionError:
                st.error("⚠️ No se pudo guardar: archivo abierto con bloqueo de escritura en Excel.")
                st.stop()
            except Exception as e:
                st.error(f"⚠️ Error al escribir direcciones: {e}")
                st.stop()